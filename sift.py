# ============================================================
# FE Portal Version 5
# 重要：fe_portal.db に従業員・シフト・時給などが保存されます。
# コード更新時に rm fe_portal.db は実行しないでください。
# バックアップは fe_portal_backups/ と users_backup.json に保存されます。
# ============================================================

from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
import sqlite3
import os
import shutil
from datetime import date, datetime, timedelta
from calendar import monthrange
from html import escape
import json
from pathlib import Path
from io import BytesIO


app = FastAPI()
DATA_DIR = Path(".")
DB = str(DATA_DIR / "fe_portal.db")
ADMIN_ID = "admin"
ADMIN_PASSWORD = "4423Kimura"
START_HOUR = 9
END_HOUR = 22
PX_PER_HOUR = 80



def db_connect():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


# =========================
# SAFE BACKUP SYSTEM
# =========================
BACKUP_DIR = DATA_DIR / "fe_portal_backups"
USER_JSON_BACKUP = DATA_DIR / "users_backup.json"
BACKUP_KEEP = 10

def table_exists(cur, table_name):
    try:
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        return cur.fetchone() is not None
    except Exception:
        return False

def count_users_in_db(db_path):
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM users")
        n = cur.fetchone()[0]
        conn.close()
        return int(n or 0)
    except Exception:
        return -1

def create_backup(reason="auto"):
    try:
        db_path = Path(DB)
        if not db_path.exists():
            print("[backup skipped] fe_portal.db がありません")
            return False

        BACKUP_DIR.mkdir(exist_ok=True)

        user_count = count_users_in_db(db_path)
        if user_count <= 0:
            print("[backup skipped] users が空、または users テーブルが読めないためバックアップしません")
            return False

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"fe_portal_{reason}_{stamp}.db"
        shutil.copy2(db_path, backup_path)

        backups = sorted(BACKUP_DIR.glob("fe_portal_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old in backups[BACKUP_KEEP:]:
            try:
                old.unlink()
            except Exception:
                pass

        save_users_json()
        print(f"[backup created] {backup_path}")
        return True
    except Exception as e:
        print(f"[backup error] {e}")
        return False

def save_users_json():
    try:
        if not Path(DB).exists():
            return False
        conn = db_connect()
        cur = conn.cursor()
        if not table_exists(cur, "users"):
            conn.close()
            return False

        cols = ["login_id", "password", "name", "is_admin"]
        for c in ["hourly_wage", "status"]:
            if has_column(cur, "users", c):
                cols.append(c)

        cur.execute("SELECT " + ",".join(cols) + " FROM users")
        rows = cur.fetchall()
        data = [dict(r) for r in rows]

        with open(USER_JSON_BACKUP, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        conn.close()
        print(f"[users json backup saved] {USER_JSON_BACKUP}")
        return True
    except Exception as e:
        print(f"[users json backup error] {e}")
        return False

def latest_backup_info():
    try:
        BACKUP_DIR.mkdir(exist_ok=True)
        backups = sorted(BACKUP_DIR.glob("fe_portal_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not backups:
            return "バックアップ未作成"
        b = backups[0]
        t = datetime.fromtimestamp(b.stat().st_mtime).strftime("%Y/%m/%d %H:%M:%S")
        return f"{b.name}（{t}）"
    except Exception:
        return "バックアップ確認エラー"

def backup_before_change(reason="change"):
    create_backup(reason)
    save_users_json()
# =========================
# END SAFE BACKUP SYSTEM
# =========================



def has_column(cur, table, col):
    cur.execute(f"PRAGMA table_info({table})")
    return any(r[1] == col for r in cur.fetchall())


def init_db():
    conn = db_connect()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        login_id TEXT UNIQUE,
        password TEXT,
        name TEXT,
        is_admin INTEGER DEFAULT 0,
        hourly_wage INTEGER DEFAULT 0
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS shifts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT,
        name TEXT,
        date TEXT,
        start TEXT,
        end TEXT,
        limit_hour TEXT,
        memo TEXT,
        confirmed INTEGER DEFAULT 0,
        cut INTEGER DEFAULT 0,
        cut_memo TEXT DEFAULT ''
    )
    """)

    # 古いDBを使っている場合でも落ちないように列を追加
    for col, ddl in [
        ("confirmed", "ALTER TABLE shifts ADD COLUMN confirmed INTEGER DEFAULT 0"),
        ("cut", "ALTER TABLE shifts ADD COLUMN cut INTEGER DEFAULT 0"),
        ("cut_memo", "ALTER TABLE shifts ADD COLUMN cut_memo TEXT DEFAULT ''"),
    ]:
        if not has_column(cur, "shifts", col):
            cur.execute(ddl)

    if not has_column(cur, "users", "hourly_wage"):
        cur.execute("ALTER TABLE users ADD COLUMN hourly_wage INTEGER DEFAULT 0")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS opinions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT,
        name TEXT,
        message TEXT,
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS published_days (
        date TEXT PRIMARY KEY,
        published INTEGER DEFAULT 0
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS day_memos (
        date TEXT PRIMARY KEY,
        memo TEXT DEFAULT ''
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS help_applications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT,
        name TEXT,
        date TEXT,
        start TEXT,
        end TEXT,
        comment TEXT DEFAULT '',
        status TEXT DEFAULT 'pending',
        created_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS labor_budgets (
        ym TEXT PRIMARY KEY,
        budget INTEGER DEFAULT 0
    )
    """)

    cur.execute("""
    INSERT OR IGNORE INTO users (login_id, password, name, is_admin)
    VALUES (?, ?, ?, 1)
    """, (ADMIN_ID, ADMIN_PASSWORD, "管理者"))


    # 従業員ステータス列を追加（社員 / オフィシャルトレーナー / アルバイト）
    try:
        cur.execute("ALTER TABLE users ADD COLUMN status TEXT DEFAULT 'アルバイト'")
    except Exception:
        pass


    # Version 5: 同日複数シフト対応のため、古い UNIQUE(user_id, date) 制約付き shifts テーブルを作り直す
    try:
        cur.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='shifts'")
        row = cur.fetchone()
        create_sql = row[0] if row else ""
        if create_sql and ("UNIQUE(user_id, date)" in create_sql or "UNIQUE(user_id,date)" in create_sql):
            cur.execute("ALTER TABLE shifts RENAME TO shifts_old_unique")
            cur.execute("""
            CREATE TABLE shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id TEXT,
                name TEXT,
                date TEXT,
                start TEXT,
                end TEXT,
                limit_hour TEXT,
                memo TEXT,
                confirmed INTEGER DEFAULT 0,
                cut INTEGER DEFAULT 0,
                cut_memo TEXT DEFAULT ''
            )
            """)
            cur.execute("""
            INSERT INTO shifts (id, user_id, name, date, start, end, limit_hour, memo, confirmed, cut, cut_memo)
            SELECT id, user_id, name, date, start, end, limit_hour, memo,
                   IFNULL(confirmed, 0), IFNULL(cut, 0), IFNULL(cut_memo, '')
            FROM shifts_old_unique
            """)
            cur.execute("DROP TABLE shifts_old_unique")
    except Exception as e:
        print("[migration warning] shifts multi migration:", e)

    # Version 5: 意見箱の未確認/確認済み管理
    try:
        if not has_column(cur, "opinions", "status"):
            cur.execute("ALTER TABLE opinions ADD COLUMN status TEXT DEFAULT 'pending'")
    except Exception:
        pass

    conn.commit()
    conn.close()


init_db()
create_backup("startup")
save_users_json()


def redirect(path):
    return RedirectResponse(path, status_code=303)


def current_user(request: Request):
    login_id = request.cookies.get("login_id")
    if not login_id:
        return None
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE login_id = ?", (login_id,))
    row = cur.fetchone()
    conn.close()
    return row


def require_login(request: Request):
    return current_user(request)


def is_admin_user(user):
    return user is not None and int(user["is_admin"] or 0) == 1


def time_options(selected=""):
    html = '<option value="">--:--</option>'
    for h in range(START_HOUR, END_HOUR + 1):
        for m in ("00", "30"):
            if h == END_HOUR and m == "30":
                continue
            t = f"{h:02d}:{m}"
            sel = "selected" if selected == t else ""
            html += f'<option value="{t}" {sel}>{t}</option>'
    return html


def limit_options(selected="指定しない"):
    values = ["指定しない", "3時間以内", "4時間以内", "5時間以内", "6時間以内", "7時間以内", "8時間以内", "9時間以内", "休み希望"]
    html = ""
    for v in values:
        sel = "selected" if selected == v else ""
        html += f'<option value="{v}" {sel}>{v}</option>'
    return html


def parse_time_to_hour(t):
    try:
        h, m = t.split(":")
        return int(h) + int(m) / 60
    except Exception:
        return None


def calc_hours(start, end):
    s = parse_time_to_hour(start)
    e = parse_time_to_hour(end)
    if s is None or e is None or e <= s:
        return 0
    return e - s


def add_month(year, month, n=1):
    month += n
    while month > 12:
        month -= 12
        year += 1
    while month < 1:
        month += 12
        year -= 1
    return year, month


def countdown_text(deadline_dt, now_dt=None):
    if now_dt is None:
        now_dt = datetime.now()

    diff_seconds = int((deadline_dt - now_dt).total_seconds())

    if diff_seconds >= 0:
        if diff_seconds <= 24 * 3600:
            hours = max(1, (diff_seconds + 3599) // 3600)
            return f'<span class="countdown-danger">残り{hours}時間！</span>'
        days = (diff_seconds + 86399) // 86400
        return f"残り{days}日"

    passed = abs(diff_seconds)
    if passed <= 24 * 3600:
        hours = max(1, (passed + 3599) // 3600)
        return f'<span class="countdown-danger">期限を{hours}時間過ぎています！</span>'
    days = (passed + 86399) // 86400
    return f'<span class="countdown-danger">期限を{days}日過ぎています！</span>'


def get_shift_period(today=None):
    now_dt = datetime.now()
    if today is None:
        today = now_dt.date()
    else:
        now_dt = datetime.combine(today, datetime.min.time())

    # 期限は毎月5日と20日。期限後3日間は同じ提出期間を表示し、4日目から次の2週間へ切り替える。
    if today.day <= 8:
        # 当月16日〜末日分を当月5日までに提出
        target_year = today.year
        target_month = today.month
        start_day = 16
        end_day = monthrange(target_year, target_month)[1]
        deadline = date(today.year, today.month, 5)
    elif today.day <= 23:
        # 翌月1日〜15日分を当月20日までに提出
        target_year, target_month = add_month(today.year, today.month, 1)
        start_day = 1
        end_day = 15
        deadline = date(today.year, today.month, 20)
    else:
        # 翌月16日〜末日分を翌月5日までに提出
        target_year, target_month = add_month(today.year, today.month, 1)
        start_day = 16
        end_day = monthrange(target_year, target_month)[1]
        deadline = date(target_year, target_month, 5)

    deadline_dt = datetime.combine(deadline, datetime.max.time()).replace(hour=23, minute=59, second=59, microsecond=0)
    remaining_text = countdown_text(deadline_dt, now_dt)
    return target_year, target_month, start_day, end_day, deadline, remaining_text


def is_published(date_value):
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT published FROM published_days WHERE date = ?", (date_value,))
    row = cur.fetchone()
    conn.close()
    return row is not None and int(row["published"] or 0) == 1


def get_day_memo(date_value):
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT memo FROM day_memos WHERE date = ?", (date_value,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return ""
    return row["memo"] or ""


def short_memo(memo, n=18):
    memo = (memo or "").strip()
    if not memo:
        return ""
    first = memo.replace("\r", "").replace("\n", " ")
    return first if len(first) <= n else first[:n] + "..."


def layout(title, body, user=None, show_nav=True, auto_scroll=True):
    user_part = ""
    if user:
        role = "管理者" if is_admin_user(user) else "スタッフ"
        user_part = f'<div class="userbar">{escape(user["name"])} さん / {role}　<a href="/logout">ログアウト</a></div>'
    nav = ""
    if show_nav and user:
        admin_link = '<a href="/admin">管理</a>' if is_admin_user(user) else '<a href="/opinion">意見箱</a>'
        nav = f"""
        <div class="bottom-space"></div>
        <nav class="footer-nav">
            <a href="/portal">ホーム</a>
            <a href="/shift">提出</a>
            <a href="/myshift">マイシフト</a>
            <a href="/shift-table">シフト表</a>
            {admin_link}
        </nav>
        """
    return f"""
    <!DOCTYPE html>
    <html lang="ja">
    <head>
        <meta charset="UTF-8">
        <title>{escape(title)}</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            :root {{
                --fe-green: #a6ce39;
                --fe-green-dark: #6a9f00;
                --fe-black: #111111;
                --fe-bg: #f7faef;
                --fe-line: #dfe9c9;
                --fe-red: #e05252;
                --fe-gray: #aeb5bd;
            }}
            * {{ box-sizing: border-box; }}
            body {{
                font-family: -apple-system, BlinkMacSystemFont, "Helvetica Neue", sans-serif;
                margin: 0;
                background: var(--fe-bg);
                color: var(--fe-black);
            }}
            header {{
                background: linear-gradient(135deg, #a6ce39 0%, #d7f26b 52%, #ffffff 100%);
                color: var(--fe-black);
                padding: 14px 18px;
                border-bottom: 4px solid var(--fe-black);
                box-shadow: 0 4px 18px rgba(0,0,0,0.18);
                position: sticky;
                top: 0;
                z-index: 30;
            }}
            .header-inner {{
                max-width: 980px;
                margin: 0 auto;
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 12px;
            }}
            .brand-box {{
                width: 42px;
                height: 42px;
                border-radius: 13px;
                background: var(--fe-black);
                color: var(--fe-green);
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 1000;
                font-size: 18px;
                box-shadow: 0 3px 0 rgba(0,0,0,0.18);
                flex: 0 0 auto;
            }}
            .header-title {{
                flex: 1;
                text-align: center;
                font-size: 23px;
                font-weight: 1000;
                letter-spacing: .04em;
                line-height: 1.15;
            }}
            .header-sub {{
                display: block;
                font-size: 11px;
                font-weight: 800;
                letter-spacing: .16em;
                opacity: .72;
                margin-top: 3px;
            }}
            .header-pill {{
                background: rgba(255,255,255,0.72);
                border: 2px solid rgba(17,17,17,0.16);
                border-radius: 999px;
                padding: 7px 10px;
                font-size: 12px;
                font-weight: 900;
                white-space: nowrap;
                flex: 0 0 auto;
            }}
            .userbar {{ background: white; padding: 10px 16px; font-size: 14px; border-bottom: 1px solid var(--fe-line); text-align: right; }}
            .userbar a {{ color: var(--fe-green-dark); font-weight: bold; }}
            .container {{ padding: 18px; max-width: 980px; margin: auto; }}
            .logo-title {{ text-align: center; font-size: 34px; font-weight: 900; margin: 30px 0 20px; }}
            .login-card, .box {{ background: white; padding: 18px; margin-bottom: 16px; border-radius: 18px; box-shadow: 0 3px 12px rgba(0,0,0,0.10); border: 1px solid var(--fe-line); }}
            .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
            .card {{ background: white; border: 2px solid var(--fe-line); border-radius: 18px; padding: 30px 10px; text-align: center; font-size: 20px; font-weight: 900; box-shadow: 0 3px 10px rgba(0,0,0,0.10); cursor: pointer; }}
            .card:hover {{ border-color: var(--fe-green); }}
            label {{ font-weight: bold; display: block; margin-top: 12px; }}
            input, select, textarea {{ padding: 12px; font-size: 16px; border: 1px solid #d7d7d7; border-radius: 12px; width: 100%; background: white; }}
            button, .btn {{ display: block; width: 100%; padding: 15px; margin-top: 14px; font-size: 18px; border: none; border-radius: 14px; background: var(--fe-green); color: var(--fe-black); font-weight: 900; text-align: center; text-decoration: none; cursor: pointer; }}
            .back {{ background: #666; color: white; }}
            .danger {{ background: var(--fe-red); color: white; }}
            .confirm {{ background: var(--fe-green-dark); color: white; }}
            .small-btn {{ display: inline-block; width: auto; padding: 8px 12px; margin: 4px 0; font-size: 14px; border-radius: 10px; }}
            .pub-on {{ background: var(--fe-green-dark) !important; color: white !important; border: 2px solid var(--fe-green-dark); }}
            .pub-off {{ background: #f2f2f2 !important; color: #555 !important; border: 2px solid #bbb; }}
            .state-on {{ color: var(--fe-green-dark); font-weight: 900; }}
            .state-off {{ color: #777; font-weight: 900; }}
            .submitted-btn {{ background: #e4f3bd; border: 2px solid var(--fe-green-dark); color: var(--fe-black); }}
            .shift-row {{ display: grid; grid-template-columns: 34px 105px 1fr 1fr 122px; gap: 8px; align-items: center; margin: 12px 0; }}
            .shift-row.submitted {{ background: #f2f8df; border-radius: 12px; padding: 6px; }}
            .date-link {{ font-size: 16px; font-weight: 900; text-decoration: none; color: #111; cursor: pointer; }}
            .date-link.submitted-date {{ text-decoration: underline; }}
            .saturday {{ color: #3977d8; }} .sunday {{ color: #e05252; }}
            input[type="checkbox"] {{ width: 24px; height: 24px; }}
            .delete-panel {{ display: none; grid-column: 1 / -1; padding: 8px 0 2px 140px; }}
            .delete-panel.show {{ display: block; }}
            .message {{ background: #e9f6c8; border-left: 6px solid var(--fe-green-dark); padding: 12px; border-radius: 12px; font-weight: bold; margin-bottom: 16px; }}
            table {{ width: 100%; border-collapse: collapse; background: white; }}
            th, td {{ border: 1px solid #ddd; padding: 10px; text-align: center; }} th {{ background: #f1f6e0; }}
            .month-form {{ display: grid; grid-template-columns: 1fr 1fr 100px; gap: 8px; align-items: end; }}
            .day-card {{ background: white; border-radius: 18px; margin-bottom: 20px; box-shadow: 0 3px 12px rgba(0,0,0,0.12); overflow: hidden; border: 2px solid var(--fe-line); }}
            .day-card.today {{ border: 3px solid var(--fe-green-dark); }}
            .day-head {{ display: grid; grid-template-columns: 90px 1fr; border-bottom: 1px solid #ddd; }}
            .day-label {{ background: #f5f5f5; text-align: center; padding: 12px 6px; font-size: 22px; font-weight: 900; border-right: 1px solid #ddd; }}
            .day-label .dow {{ font-size: 26px; }} .day-label .date-num {{ font-size: 28px; margin-top: 8px; }}
            .day-memo-chip {{
                margin-top: 8px;
                font-size: 13px;
                line-height: 1.25;
                color: #555;
                font-weight: 900;
                cursor: pointer;
                word-break: break-all;
            }}
            .day-memo-chip:hover {{ color: var(--fe-green-dark); }}
            .day-publish-control {{
                position: absolute;
                right: 10px;
                top: 34px;
                z-index: 8;
                display: flex;
                gap: 6px;
                align-items: center;
                background: rgba(255,255,255,0.92);
                border: 1px solid var(--fe-line);
                border-radius: 12px;
                padding: 6px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            }}
            .day-publish-control .publish-state {{ font-size: 12px; min-width: 56px; }}
            .day-publish-control .publish-btn {{ width: 78px; min-width: 78px; font-size: 12px; padding: 6px 8px; margin: 0; }}
            .memo-modal {{
                display: none;
                position: fixed;
                inset: 0;
                background: rgba(0,0,0,0.45);
                z-index: 999;
                align-items: center;
                justify-content: center;
                padding: 24px;
            }}
            .memo-modal.show {{ display: flex; }}
            .memo-box {{
                width: min(520px, 92vw);
                background: white;
                border-radius: 22px;
                padding: 24px;
                box-shadow: 0 12px 36px rgba(0,0,0,0.28);
            }}
            .memo-title-row {{ display:flex; justify-content:space-between; align-items:center; gap:16px; }}
            .memo-close-x {{ font-size: 30px; font-weight: 900; cursor: pointer; }}
            .memo-body {{ white-space: pre-wrap; line-height: 1.8; font-weight: 800; color:#555; font-size: 17px; margin: 26px 0; }}
            .timeline-wrap {{ overflow-x: auto; }}
            .timeline {{ position: relative; min-width: {((END_HOUR - START_HOUR) * PX_PER_HOUR) + 100}px; min-height: 190px; background: white; }}
            .time-line {{ position: absolute; top: 0; bottom: 0; width: 1px; background: #cfcfcf; }}
            .time-label {{ position: absolute; top: 6px; font-size: 13px; color: #666; }}
            .bar {{ position: absolute; height: 30px; border-radius: 8px; color: white; font-size: 14px; font-weight: bold; padding: 0 8px; line-height: 30px; overflow: hidden; white-space: nowrap; box-shadow: inset 0 -1px 0 rgba(0,0,0,0.15); text-decoration: none; display: block; }}
            .bar.pending {{ opacity: 0.45; border: 2px dashed rgba(0,0,0,0.35); }}
            .bar.gray {{ background: var(--fe-gray); color: transparent; }}
            .bar.cut {{ opacity: 0.7; text-decoration: line-through; color: #111; background: #ddd !important; border: 2px solid var(--fe-red); }}
            .bar.help {{ background: #ee5b5f !important; color: white; border: 2px solid #c82333; }}
            .bar.help-slot {{ background: #ee5b5f !important; color: white; border: 2px solid #c82333; text-align:center; }}
            .bar.help-app {{ background: #f5a623 !important; color: white; border: 2px solid #d98200; }}
            .help-note {{ background:#fff2d6; border-left:6px solid #f5a623; padding:10px; border-radius:12px; margin-bottom:12px; font-weight:bold; }}
            .action-panel {{ display: none; position: absolute; z-index: 9; background: white; border: 2px solid var(--fe-green-dark); border-radius: 12px; padding: 8px; box-shadow: 0 4px 16px rgba(0,0,0,0.2); min-width: 150px; }}
            .action-panel.show {{ display: block; }}
            .action-panel .btn {{ margin-top: 6px; padding: 8px 10px; font-size: 13px; border-radius: 8px; }}
            .lock-mark {{ color: var(--fe-red); font-size: 26px; margin-top: 8px; }}
            .empty-note {{ position: absolute; left: 12px; top: 60px; color: #999; font-weight: bold; }}
            .summary {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
            .summary .box {{ text-align: center; }} .summary-num {{ font-size: 28px; font-weight: 900; color: var(--fe-green-dark); }}
            .countdown-danger {{ color: #d60000; font-weight: 900; }}
            .status-ok {{ color: var(--fe-green-dark); font-weight: 900; }}
            .status-ng {{ color: #d60000; font-weight: 900; }}
            .calendar {{ display: grid; grid-template-columns: repeat(7, 1fr); gap: 4px; margin-bottom: 18px; }}
            .cal-head {{ text-align: center; font-weight: 900; padding: 8px 0; color: #666; }}
            .cal-day {{ min-height: 74px; background: white; border: 1px solid var(--fe-line); border-radius: 10px; padding: 6px; font-size: 13px; }}
            .cal-day.today {{ border: 3px solid var(--fe-green-dark); }}
            .cal-date {{ font-weight: 900; margin-bottom: 4px; }}
            .cal-shift {{ background: #e4f3bd; color: #111; border-radius: 8px; padding: 3px; font-size: 12px; font-weight: 800; }}
            .employee-row {{ display: grid; grid-template-columns: 1fr auto auto; gap: 8px; align-items: center; }}
            .footer-nav {{ position: fixed; left: 0; right: 0; bottom: 0; background: white; border-top: 1px solid #ddd; display: grid; grid-template-columns: repeat(5, 1fr); z-index: 20; }}
            .footer-nav a {{ text-align: center; padding: 10px 4px; color: #111; text-decoration: none; font-weight: bold; font-size: 13px; }}
            .bottom-space {{ height: 56px; }}

            .publish-table {{
                table-layout: fixed;
                width: 100%;
            }}

            .publish-table th:nth-child(1),
            .publish-table td:nth-child(1) {{
                width: 20%;
            }}

            .publish-table th:nth-child(2),
            .publish-table td:nth-child(2) {{
                width: 18%;
            }}

            .publish-table th:nth-child(3),
            .publish-table td:nth-child(3) {{
                width: 24%;
            }}

            .publish-table th:nth-child(4),
            .publish-table td:nth-child(4) {{
                width: 38%;
            }}

            .publish-state {{
                display: inline-block;
                min-width: 72px;
                text-align: center;
                font-weight: 900;
            }}

            .publish-btn {{
                width: 128px;
                min-width: 128px;
                text-align: center;
                white-space: nowrap;
            }}

            @media screen and (max-width: 650px) {{
                .container {{ padding: 14px; }}
                .shift-row {{ grid-template-columns: 30px 85px 1fr 1fr 92px; gap: 6px; }}
                input, select {{ font-size: 14px; padding: 10px 6px; }}
                .date-link {{ font-size: 14px; }}
                .month-form {{ grid-template-columns: 1fr 1fr; }} .month-form button {{ grid-column: 1 / -1; }}
                .day-head {{ grid-template-columns: 80px 1fr; }}
                .calendar {{ gap: 3px; }} .cal-day {{ min-height: 64px; padding: 4px; font-size: 12px; }}
                header {{ padding: 12px 10px; }}
                .brand-box {{ width: 36px; height: 36px; font-size: 15px; border-radius: 11px; }}
                .header-title {{ font-size: 20px; }}
                .header-pill {{ font-size: 10px; padding: 6px 8px; }}
            }}
        
            .badge-card {{
                position: relative;
            }}
            .app-badge {{
                position: absolute;
                top: 8px;
                right: 10px;
                min-width: 26px;
                height: 26px;
                padding: 0 8px;
                border-radius: 999px;
                background: #e60012;
                color: #fff;
                font-size: 14px;
                line-height: 26px;
                font-weight: 1000;
                box-shadow: 0 2px 8px rgba(0,0,0,.22);
            }}
            .notify-box {{
                border: 2px solid #e60012 !important;
                color: #e60012;
                font-weight: 900;
            }}

        
            .badge-card {{
                position: relative;
            }}
            .app-badge {{
                position: absolute;
                top: 8px;
                right: 10px;
                min-width: 26px;
                height: 26px;
                padding: 0 8px;
                border-radius: 999px;
                background: #e60012;
                color: #fff;
                font-size: 14px;
                line-height: 26px;
                font-weight: 1000;
                box-shadow: 0 2px 8px rgba(0,0,0,.22);
            }}
            .notify-box {{
                border: 2px solid #e60012 !important;
                color: #e60012;
                font-weight: 900;
            }}
            .help-status-pending {{
                color: #e60012;
                font-weight: 1000;
            }}
            .help-status-approved {{
                color: #6a9f00;
                font-weight: 1000;
            }}
            .help-status-rejected {{
                color: #777;
                font-weight: 900;
            }}

        
            .bar.employee-bar {{
                background: #111111 !important;
                color: #ffffff !important;
                font-weight: 1000;
                border: 1px solid #000;
            }}

        </style>
    </head>
    <body>
        <header>
            <div class="header-inner">
                <div class="brand-box">FE</div>
                <div class="header-title">{escape(title)}<span class="header-sub">FIT-EASY PORTAL</span></div>
                <div class="header-pill">Portal</div>
            </div>
        </header>
        {user_part}
        <div class="container">{body}</div>
        <div id="dayMemoModal" class="memo-modal" onclick="if(event.target.id==='dayMemoModal') closeDayMemo()">
            <div class="memo-box">
                <div class="memo-title-row">
                    <h2 id="dayMemoTitle">メモ</h2>
                    <div class="memo-close-x" onclick="closeDayMemo()">×</div>
                </div>
                <div id="dayMemoBody" class="memo-body"></div>
                <button type="button" class="back" onclick="closeDayMemo()">閉じる</button>
            </div>
        </div>
        {nav}
        <script>
            function toggleDelete(id) {{
                const el = document.getElementById(id);
                if (el) el.classList.toggle('show');
            }}
            function toggleAction(id) {{
                document.querySelectorAll('.action-panel').forEach(function(panel) {{
                    if (panel.id !== id) panel.classList.remove('show');
                }});
                const el = document.getElementById(id);
                if (el) el.classList.toggle('show');
            }}
            async function setPublishAjax(day, value) {{
                try {{
                    const res = await fetch('/set-publish/' + day + '/' + value + '?ajax=1');
                    if (!res.ok) throw new Error('通信エラー');
                    const nextValue = value === 1 ? 0 : 1;
                    const state = document.getElementById('pub-state-' + day);
                    const btn = document.getElementById('pub-btn-' + day);
                    if (state) {{
                        state.textContent = value === 1 ? '公開中' : '非公開';
                        state.className = value === 1 ? 'publish-state state-on' : 'publish-state state-off';
                    }}
                    if (btn) {{
                        btn.textContent = value === 1 ? '非公開にする' : '公開する';
                        btn.className = value === 1 ? 'btn small-btn publish-btn pub-off' : 'btn small-btn publish-btn pub-on';
                        btn.setAttribute('onclick', "setPublishAjax('" + day + "', " + nextValue + "); return false;");
                        btn.href = '/set-publish/' + day + '/' + nextValue;
                    }}
                }} catch(e) {{
                    alert('公開設定の変更に失敗しました');
                }}
            }}
            function showDayMemo(title, memo) {{
                const modal = document.getElementById('dayMemoModal');
                const titleEl = document.getElementById('dayMemoTitle');
                const bodyEl = document.getElementById('dayMemoBody');
                if (!modal || !titleEl || !bodyEl) return;
                titleEl.textContent = title;
                bodyEl.textContent = memo;
                modal.classList.add('show');
            }}
            function closeDayMemo() {{
                const modal = document.getElementById('dayMemoModal');
                if (modal) modal.classList.remove('show');
            }}
            const AUTO_SCROLL_TO_TODAY = {str(auto_scroll).lower()};
            window.addEventListener('load', function() {{
                const path = window.location.pathname;
                if (path === '/admin-shifts') {{
                    const y = sessionStorage.getItem('adminShiftsScrollY');
                    if (y !== null) {{
                        window.scrollTo(0, parseInt(y || '0', 10));
                        sessionStorage.removeItem('adminShiftsScrollY');
                        return;
                    }}
                }}
                if (!AUTO_SCROLL_TO_TODAY) return;
                const today = document.querySelector('[data-today="1"]');
                if (today) today.scrollIntoView({{behavior: 'smooth', block: 'start'}});
            }});

            window.addEventListener('click', function(e) {{
                const a = e.target.closest('a');
                if (!a) return;
                const href = a.getAttribute('href') || '';
                if (window.location.pathname === '/admin-shifts' &&
                    (href.includes('/admin-shift-action/') || href.includes('/confirm-shift/') || href.includes('/cut-shift/'))) {{
                    sessionStorage.setItem('adminShiftsScrollY', String(window.scrollY));
                }}
            }});
        </script>
    </body>
    </html>
    """


def login_page(message=""):
    msg = f'<div class="message">{escape(message)}</div>' if message else ""
    body = f"""
    <div class="logo-title">FE Portal</div>
    {msg}
    <div class="login-card">
        <form action="/login" method="post">
            <label>ID</label><input name="login_id" placeholder="ID" required>
            <label>パスワード</label><input type="password" name="password" placeholder="パスワード" required>
            <button type="submit">ログイン</button>
        </form>
        <a class="btn back" href="/register">新規登録</a>
    </div>
    <div class="box">管理者ログイン：ID <b>admin</b> / パスワード <b>4423Kimura</b></div>
    """
    return layout("FE Portal ログイン", body, user=None, show_nav=False)


@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    user = current_user(request)
    if user:
        return redirect("/portal")
    return login_page()


@app.post("/login")
def login(login_id: str = Form(""), password: str = Form("")):
    login_id = login_id.strip()
    password = password.strip()
    if not login_id:
        return HTMLResponse(login_page("IDを入力してください。"))
    if not password:
        return HTMLResponse(login_page("パスワードを入力してください。"))
    conn = db_connect(); cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE login_id = ? AND password = ?", (login_id, password))
    user = cur.fetchone(); conn.close()
    if not user:
        return HTMLResponse(login_page("IDまたはパスワードが違います。"))
    res = redirect("/portal")
    res.set_cookie("login_id", login_id, max_age=60 * 60 * 24 * 30)
    return res



@app.get("/settings", response_class=HTMLResponse)
def settings(request: Request):
    user=require_login(request)
    if not user:
        return redirect("/")
    body=f"""
    <h2>設定</h2>
    <div class='box'>
    氏名：{escape(user['name'])}<br>
    ID：{escape(user['login_id'])}<br>
    ステータス：{escape(user.get('status','')) if hasattr(user,'get') else ''}
    </div>
    <form action='/change-password' method='post'>
    <label>現在のパスワード</label><input type='password' name='current_password' required>
    <label>新しいパスワード</label><input type='password' name='new_password' required>
    <label>新しいパスワード（確認）</label><input type='password' name='confirm_password' required>
    <button type='submit'>パスワード変更</button>
    </form>
    <a class='btn back' href='/portal'>戻る</a>
    """
    return layout("設定", body, user=user)

@app.post("/change-password")
def change_password(request: Request, current_password: str = Form(...), new_password: str = Form(...), confirm_password: str = Form(...)):
    user=require_login(request)
    if not user:
        return redirect('/')
    if new_password != confirm_password:
        return HTMLResponse("<script>alert('新しいパスワードが一致しません');history.back();</script>")
    conn=db_connect(); cur=conn.cursor()
    cur.execute('SELECT * FROM users WHERE login_id=?',(user['login_id'],))
    row=cur.fetchone()
    if not row or row['password'] != current_password:
        conn.close()
        return HTMLResponse("<script>alert('現在のパスワードが違います');history.back();</script>")
    cur.execute('UPDATE users SET password=? WHERE login_id=?',(new_password,user['login_id']))
    conn.commit(); conn.close()
    return HTMLResponse("<script>alert('パスワードを変更しました');window.location='/portal';</script>")

@app.get("/logout")
def logout():
    res = redirect("/")
    res.delete_cookie("login_id")
    return res


@app.get("/register", response_class=HTMLResponse)
def register_page():
    body = """
    <div class="logo-title">新規登録</div>
    <div class="login-card">
        <form action="/register" method="post">
            <label>名前</label><input name="name" placeholder="例：木村洸介" required>
            <label>ID</label><input name="login_id" placeholder="ログインに使うID" required>
            <label>パスワード</label><input type="password" name="password" placeholder="パスワード" required>
            <button type="submit">登録する</button>
        </form>
        <a class="btn back" href="/">ログイン画面へ戻る</a>
    </div>
    """
    return layout("新規登録", body, show_nav=False)


@app.post("/register")
def register(name: str = Form(""), login_id: str = Form(""), password: str = Form("")):
    name = name.strip()
    login_id = login_id.strip()
    password = password.strip()
    if not name:
        return HTMLResponse(login_page("名前を入力してください。"))
    if not login_id:
        return HTMLResponse(login_page("IDを入力してください。"))
    if not password:
        return HTMLResponse(login_page("パスワードを入力してください。"))
    conn = db_connect(); cur = conn.cursor()
    try:
        cur.execute("INSERT INTO users (login_id, password, name, is_admin) VALUES (?, ?, ?, 0)", (login_id, password, name))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return HTMLResponse(login_page("そのIDはすでに使われています。"))
    conn.close()
    res = redirect("/portal")
    res.set_cookie("login_id", login_id, max_age=60 * 60 * 24 * 30)
    return res


@app.get("/portal", response_class=HTMLResponse)
def portal(request: Request):
    user = require_login(request)
    if not user:
        return redirect("/")
    admin_card = '<div class="card" onclick="location.href=\'/admin\'">管理画面</div>' if is_admin_user(user) else '<div class="card" onclick="location.href=\'/opinion\'">意見箱</div>'
    body = f"""
    <h2>FEポータル</h2>
    <div class="grid">
        <div class="card" onclick="location.href='/shift'">シフト提出</div>
        <div class="card" onclick="location.href='/myshift'">マイシフト</div>
        <div class="card" onclick="location.href='/shift-table'">シフト表</div>
        <div class="card" onclick="location.href='/salary'">給与明細</div>
        {admin_card}
        <div class="card" onclick="location.href='/settings'">設定</div>
        <div class="card" onclick="location.href='/logout'">ログアウト</div>
    </div>
    """
    return layout("FEポータル", body, user=user)


@app.get("/shift", response_class=HTMLResponse)
def shift(request: Request, saved: int = 0):
    user = require_login(request)
    if not user:
        return redirect("/")
    year, month, start_day, end_day, deadline, remaining_text = get_shift_period()
    start_date = f"{year}-{month:02d}-{start_day:02d}"
    end_date = f"{year}-{month:02d}-{end_day:02d}"
    conn = db_connect(); cur = conn.cursor()
    cur.execute("SELECT * FROM shifts WHERE user_id = ? AND date BETWEEN ? AND ?", (user["login_id"], start_date, end_date))
    saved_rows = {r["date"]: r for r in cur.fetchall()}
    conn.close()
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    rows = ""
    for d in range(start_day, end_day + 1):
        this_date = date(year, month, d)
        youbi = weekdays[this_date.weekday()]
        date_label = f"{month:02d}/{d:02d}({youbi})"
        date_value = f"{year}-{month:02d}-{d:02d}"
        r = saved_rows.get(date_value)
        submitted = r is not None
        day_class = "saturday" if youbi == "土" else "sunday" if youbi == "日" else ""
        checked = "checked" if submitted else ""
        submitted_class = "submitted" if submitted else ""
        date_class = "submitted-date" if submitted else ""
        delete_panel = ""
        if submitted:
            delete_panel = f"""
            <div id="delete-{date_value}" class="delete-panel">
                <a class="btn danger small-btn" href="/delete-my-shift/{date_value}">削除</a>
            </div>
            """
        rows += f"""
        <div class="shift-row {submitted_class}">
            <input type="checkbox" name="selected_dates" value="{date_value}" {checked}>
            <a class="date-link {day_class} {date_class}" onclick="toggleDelete('delete-{date_value}')">{date_label}</a>
            <select name="start_{date_value}">{time_options(r['start'] if submitted else '')}</select>
            <select name="end_{date_value}">{time_options(r['end'] if submitted else '')}</select>
            <select name="limit_{date_value}">{limit_options(r['limit_hour'] if submitted else '指定しない')}</select>
            {delete_panel}
        </div>
        """
    msg = '<div class="message">シフト提出済み</div>' if saved else ""
    button_class = "submitted-btn" if saved_rows else ""
    button_text = "提出済み" if saved_rows else "提出する"
    body = f"""
    <h2>シフト提出</h2>
    {msg}
    <div class="box"><div>提出締切</div><div class="deadline">{deadline.strftime('%Y/%m/%d')} 23:59</div><div>残り時間</div><div class="summary-num">{remaining_text}</div></div>
    <div class="box"><b>提出対象</b><br>{year}年{month}月{start_day}日〜{end_day}日</div>
    <form action="/shift-submit" method="post">
        <label>名前</label><input value="{escape(user['name'])}" readonly style="background:#f0f0f0; color:#555;">
        <br><br>{rows}
        <label>メモ</label><textarea name="memo" placeholder="終電、テスト期間など"></textarea>
        <button class="{button_class}" type="submit">{button_text}</button>
    </form>
    <a class="btn back" href="/portal">戻る</a>
    """
    return layout("シフト提出", body, user=user)


@app.post("/shift-submit")
async def shift_submit(request: Request):
    user = require_login(request)
    if not user:
        return redirect("/")
    form = await request.form()
    name = user["name"]
    memo = form.get("memo", "")
    selected_dates = set(form.getlist("selected_dates"))
    year, month, start_day, end_day, _, _ = get_shift_period()
    conn = db_connect(); cur = conn.cursor()
    for d in range(start_day, end_day + 1):
        date_value = f"{year}-{month:02d}-{d:02d}"
        start = form.get(f"start_{date_value}", "")
        end = form.get(f"end_{date_value}", "")
        limit_hour = form.get(f"limit_{date_value}", "指定しない")
        should_save = date_value in selected_dates or start or end or limit_hour not in ("", "指定しない")

        # 通常提出分だけ更新する。管理者追加・ヘルプ承認などの確定シフトは消さない。
        cur.execute("""
        DELETE FROM shifts
        WHERE user_id = ? AND date = ?
          AND IFNULL(confirmed, 0) = 0
          AND IFNULL(limit_hour, '') NOT IN ('管理者追加', 'ヘルプ承認')
        """, (user["login_id"], date_value))

        if should_save:
            cur.execute("""
            INSERT INTO shifts (user_id, name, date, start, end, limit_hour, memo, confirmed, cut, cut_memo)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0, 0, '')
            """, (user["login_id"], name, date_value, start, end, limit_hour, memo))
    conn.commit(); conn.close()
    return redirect("/shift?saved=1")


@app.get("/delete-my-shift/{date_value}")
def delete_my_shift(date_value: str, request: Request):
    user = require_login(request)
    if not user:
        return redirect("/")
    conn = db_connect(); cur = conn.cursor()
    cur.execute("""
    DELETE FROM shifts
    WHERE user_id = ? AND date = ?
      AND IFNULL(confirmed, 0) = 0
      AND IFNULL(limit_hour, '') NOT IN ('管理者追加', 'ヘルプ承認')
    """, (user["login_id"], date_value))
    conn.commit(); conn.close()
    return redirect("/shift")


def calendar_html(year, month, rows):
    # マイシフト上部のカレンダーには、管理者が確定したシフトだけ表示する。
    # 同じ日に複数シフトがある場合は複数行で表示する。
    shifts = {}
    for r in rows:
        if int(r["cut"] or 0) == 0 and int(r["confirmed"] or 0) == 1:
            shifts.setdefault(r["date"], []).append(r)

    first_wd = date(year, month, 1).weekday()  # 月=0
    last_day = monthrange(year, month)[1]
    heads = ["月", "火", "水", "木", "土", "日"]
    heads = ["月", "火", "水", "木", "金", "土", "日"]
    html = '<div class="calendar">'
    for h in heads:
        html += f'<div class="cal-head">{h}</div>'
    for _ in range(first_wd):
        html += '<div class="cal-day" style="opacity:.35"></div>'
    today_iso = date.today().isoformat()
    for d in range(1, last_day + 1):
        day = f"{year}-{month:02d}-{d:02d}"
        cls = " today" if day == today_iso else ""
        item = ""
        if day in shifts:
            for r in shifts[day]:
                item += f'<div class="cal-shift">{escape(r["start"] or "--:--")}-{escape(r["end"] or "--:--")}</div>'
        html += f'<div class="cal-day{cls}"><div class="cal-date">{d}</div>{item}</div>'
    html += '</div>'
    return html


@app.get("/myshift", response_class=HTMLResponse)
def myshift(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user:
        return redirect("/")
    today = date.today(); year = year or today.year; month = month or today.month
    start_month = f"{year}-{month:02d}-01"; end_month = f"{year}-{month:02d}-{monthrange(year, month)[1]:02d}"
    conn = db_connect(); cur = conn.cursor()
    cur.execute("SELECT * FROM shifts WHERE user_id = ? AND date BETWEEN ? AND ? ORDER BY date ASC", (user["login_id"], start_month, end_month))
    rows = cur.fetchall(); conn.close()
    confirmed_rows = [r for r in rows if int(r["confirmed"] or 0) == 1 and int(r["cut"] or 0) == 0]
    total = sum(calc_hours(r["start"], r["end"]) for r in confirmed_rows)
    body = f"""
    <h2>マイシフト</h2>
    <form class="month-form" action="/myshift" method="get">
        <div><label>年</label><input type="number" name="year" value="{year}"></div>
        <div><label>月</label><input type="number" name="month" value="{month}"></div>
        <button type="submit">表示</button>
    </form>
    {calendar_html(year, month, confirmed_rows)}
    <div class="summary"><div class="box"><div>提出回数</div><div class="summary-num">{len(rows)}</div></div><div class="box"><div>確定シフト合計時間</div><div class="summary-num">{total:.1f}</div></div></div>
    """
    if not rows:
        body += '<div class="box">この月のシフトはまだありません。</div>'
    for r in rows:
        state = "確定" if int(r["confirmed"] or 0) == 1 else "削り" if int(r["cut"] or 0) == 1 else "提出済み"
        body += f"""
        <div class="box">
            <b>{escape(r['date'])}</b><br>
            {escape(r['start'] or '--:--')} - {escape(r['end'] or '--:--')}<br>
            希望：{escape(r['limit_hour'] or '指定しない')}<br>
            状態：{state}<br>
            メモ：{escape(r['memo'] or '')}
        </div>
        """
    body += '<a class="btn back" href="/portal">戻る</a>'
    return layout("マイシフト", body, user=user)


def help_segments(shifts):
    """提出シフトから、人手不足の時間帯を30分単位でまとめる。
    9:00-10:00 と 21:00-22:00 は最低3人、
    それ以外の30分枠は最低1人として不足人数を出す。
    """
    slots = []
    t = START_HOUR
    while t < END_HOUR:
        is_opening_hour = abs(t - START_HOUR) < 0.001
        is_closing_hour = abs(t - (END_HOUR - 1)) < 0.001
        requirement = 3 if (is_opening_hour or is_closing_hour) else 1
        count = 0
        for r in shifts:
            if int(r["cut"] or 0) == 1:
                continue
            st = parse_time_to_hour(r["start"])
            en = parse_time_to_hour(r["end"])
            if st is None or en is None:
                continue
            if st <= t < en:
                count += 1
        deficit = requirement - count
        if deficit > 0:
            slots.append((t, deficit))
        t += 0.5

    segments = []
    if not slots:
        return segments

    seg_start, seg_deficit = slots[0]
    prev_t = seg_start

    for t, deficit in slots[1:]:
        if abs(t - (prev_t + 0.5)) < 0.001 and deficit == seg_deficit:
            prev_t = t
        else:
            segments.append((seg_start, prev_t + 0.5, seg_deficit))
            seg_start, seg_deficit = t, deficit
            prev_t = t

    segments.append((seg_start, prev_t + 0.5, seg_deficit))
    return segments




def get_user_status(login_id):
    try:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT status FROM users WHERE login_id = ?", (login_id,))
        row = cur.fetchone()
        conn.close()
        if row and "status" in row.keys() and row["status"]:
            return row["status"]
    except Exception:
        pass
    return "アルバイト"

def get_help_applications(start_date, end_date, include_all=False):
    conn = db_connect()
    cur = conn.cursor()
    if include_all:
        cur.execute("""
        SELECT * FROM help_applications
        WHERE date BETWEEN ? AND ?
        ORDER BY date ASC, start ASC, id ASC
        """, (start_date, end_date))
    else:
        cur.execute("""
        SELECT * FROM help_applications
        WHERE date BETWEEN ? AND ? AND status = 'pending'
        ORDER BY date ASC, start ASC, id ASC
        """, (start_date, end_date))
    rows = cur.fetchall()
    conn.close()
    grouped = {}
    for r in rows:
        grouped.setdefault(r["date"], []).append(r)
    return grouped


def fmt_hour(h):
    hh = int(h)
    mm = "30" if abs(h - hh - 0.5) < 0.01 else "00"
    return f"{hh:02d}:{mm}"


def hour_to_str(h):
    return fmt_hour(h)

def timeline_html(day, shifts, published, admin=False, manage=False, year=None, month=None, help_apps=None, viewer=None):
    help_apps = help_apps or []
    dt = date.fromisoformat(day)
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    today_flag = "1" if day == date.today().isoformat() else "0"
    today_class = "today" if today_flag == "1" else ""
    lock = "" if published or admin or manage else '<div class="lock-mark">🔒</div>'
    memo = get_day_memo(day)
    memo_chip = ""
    if memo:
        memo_title = f"{dt.month}/{dt.day} メモ"
        memo_title_js = json.dumps(memo_title, ensure_ascii=False)
        memo_js = json.dumps(memo, ensure_ascii=False)
        memo_chip = (
            f'<div class="day-memo-chip" onclick="showDayMemo({memo_title_js}, {memo_js})">'
            f'📌{escape(short_memo(memo, 16))}</div>'
        )
    html = f"""
    <div id="day-{day}" class="day-card {today_class}" data-today="{today_flag}">
        <div class="day-head">
            <div class="day-label"><div class="dow">{weekdays[dt.weekday()]}</div><div class="date-num">{dt.day}</div>{lock}{memo_chip}</div>
            <div class="timeline-wrap"><div class="timeline">
    """
    for h in range(START_HOUR, END_HOUR + 1):
        left = (h - START_HOUR) * PX_PER_HOUR
        html += f'<div class="time-line" style="left:{left}px;"></div><div class="time-label" style="left:{left + 4}px;">{h}</div>'

    if manage:
        pub = is_published(day)
        state = "公開中" if pub else "非公開"
        val = 0 if pub else 1
        action = "非公開" if pub else "公開"
        state_class = "publish-state state-on" if pub else "publish-state state-off"
        btn_class = "pub-off" if pub else "pub-on"
        add_url = f"/admin-shift-new?year={year}&month={month}&day={day}"
        html += f"""
        <div class="day-publish-control">
            <span id="pub-state-{day}" class="{state_class}">{state}</span>
            <a id="pub-btn-{day}" class="btn small-btn publish-btn {btn_class}" href="/set-publish/{day}/{val}" onclick="setPublishAjax('{day}', {val}); return false;">{action}</a>
            <a class="btn small-btn confirm" href="{add_url}">＋追加</a>
        </div>
        """

    visible = (published or admin or manage)
    if not visible:
        html += '<div class="empty-note">非公開</div>'
    elif not shifts:
        html += '<div class="empty-note">シフトなし</div>'

    colors = ["#8ac053", "#4fa3a5", "#6b7fd7", "#9b65c9", "#d6b936", "#333333"]
    y = 36

    if visible:
        sorted_shifts = sorted(
            shifts,
            key=lambda r: (
                1 if get_user_status(r["user_id"]) == "社員" else 0,
                parse_time_to_hour(r["start"]) if parse_time_to_hour(r["start"]) is not None else 999,
                r["name"] or ""
            )
        )
        for i, r in enumerate(sorted_shifts):
            emp_status = get_user_status(r["user_id"])
            start_h = parse_time_to_hour(r["start"])
            end_h = parse_time_to_hour(r["end"])
            if start_h is None or end_h is None or end_h <= start_h:
                continue
            start_h = max(start_h, START_HOUR)
            end_h = min(end_h, END_HOUR)
            left = (start_h - START_HOUR) * PX_PER_HOUR
            width = max((end_h - start_h) * PX_PER_HOUR, 42)
            color = "#111111" if emp_status == "社員" else colors[i % len(colors)]
            text = f"{escape(r['name'])} {calc_hours(r['start'], r['end']):.1f}"
            if int(r["cut"] or 0) == 1:
                bar_state_class = " cut"
            elif int(r["confirmed"] or 0) == 1:
                bar_state_class = ""
            else:
                bar_state_class = " pending"

            employee_class = " employee-bar" if emp_status == "社員" else ""

            if manage:
                panel_id = f"action-{r['id']}"
                q = f"?year={year}&month={month}" if year and month else ""
                html += f"""
                <a class="bar{bar_state_class}{employee_class}" href="javascript:void(0)" onclick="toggleAction('{panel_id}')" style="left:{left}px; top:{y}px; width:{width}px; background:{color};">{text}</a>
                <div id="{panel_id}" class="action-panel" style="left:{left}px; top:{y + 32}px;">
                    <b>{escape(r['name'])}</b><br>
                    {escape(r['start'] or '--:--')} - {escape(r['end'] or '--:--')}<br>
                    <a class="btn confirm" href="/confirm-shift/{r['id']}{q}">確定</a>
                    <a class="btn danger" href="/cut-shift/{r['id']}{q}">削る</a>
                    <a class="btn danger" href="/delete-shift-admin/{r['id']}{q}" onclick="return confirm('このシフトを完全に削除します。よろしいですか？');">削除</a>
                </div>
                """
            else:
                html += f'<div class="bar{bar_state_class}{employee_class}" style="left:{left}px; top:{y}px; width:{width}px; background:{color};">{text}</div>'
            y += 34

        # ヘルプ応募中バーを表示
        app_y = max(y + 8, 138)
        for app in help_apps:
            st = parse_time_to_hour(app["start"])
            en = parse_time_to_hour(app["end"])
            if st is None or en is None or en <= st:
                continue
            st = max(st, START_HOUR)
            en = min(en, END_HOUR)
            left = (st - START_HOUR) * PX_PER_HOUR
            width = max((en - st) * PX_PER_HOUR, 80)
            app_text = f"応募：{escape(app['name'])}"
            if manage or admin:
                html += f'<a class="bar help-app" href="/admin-help-action/{app["id"]}?year={year}&month={month}" style="left:{left}px; top:{app_y}px; width:{width}px;">{app_text}</a>'
            else:
                if viewer and app["user_id"] == viewer["login_id"]:
                    html += f'<a class="bar help-app" href="/cancel-help/{app["id"]}" style="left:{left}px; top:{app_y}px; width:{width}px;">応募中 取消</a>'
                else:
                    html += f'<div class="bar help-app" style="left:{left}px; top:{app_y}px; width:{width}px;">応募あり</div>'
            app_y += 34

        # 不足人数ぶん、赤いヘルプバーを複数本表示
        help_y = max(app_y + 8, 138)
        for hs, he, deficit in help_segments(shifts):
            left = (hs - START_HOUR) * PX_PER_HOUR
            width = max((he - hs) * PX_PER_HOUR, 70)
            for n in range(deficit):
                label = "ヘルプ"
                if manage or admin:
                    html += f'<div class="bar help-slot" style="left:{left}px; top:{help_y}px; width:{width}px;">{label}</div>'
                else:
                    start_s = fmt_hour(hs)
                    end_s = fmt_hour(he)
                    html += f'<a class="bar help-slot" href="/help-apply?year={year}&month={month}&date={day}&start={start_s}&end={end_s}&slot={n+1}" style="left:{left}px; top:{help_y}px; width:{width}px;">{label}</a>'
                help_y += 34
        if help_y > 190:
            html += f'<style>.day-card[data-today="{today_flag}"] .timeline {{ min-height: {help_y + 40}px; }}</style>'

    html += "</div></div></div></div>"
    return html


def build_pass_rate_table(start_date, end_date):
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
    SELECT name, confirmed, cut
    FROM shifts
    WHERE date BETWEEN ? AND ?
    ORDER BY name ASC
    """, (start_date, end_date))
    rows = cur.fetchall()
    conn.close()

    stats = {}
    for r in rows:
        name = r["name"] or "未登録"
        if name not in stats:
            stats[name] = {"total": 0, "passed": 0, "cut": 0, "pending": 0}
        stats[name]["total"] += 1
        if int(r["cut"] or 0) == 1:
            stats[name]["cut"] += 1
        elif int(r["confirmed"] or 0) == 1:
            stats[name]["passed"] += 1
        else:
            stats[name]["pending"] += 1

    html = """
    <div class="box">
        <h3>希望通過率</h3>
        <p class="muted">確定したシフト ÷ 提出したシフトで計算</p>
        <table>
            <tr>
                <th>名前</th>
                <th>希望通過率</th>
                <th>確定</th>
                <th>削り</th>
                <th>未確定</th>
            </tr>
    """

    if not stats:
        html += """
            <tr><td colspan="5">まだ提出シフトがありません。</td></tr>
        """
    else:
        for name, st in stats.items():
            total = st["total"]
            rate = round(st["passed"] / total * 100) if total else 0
            html += f"""
            <tr>
                <td>{escape(name)}</td>
                <td><b>{rate}%</b></td>
                <td>{st['passed']}</td>
                <td>{st['cut']}</td>
                <td>{st['pending']}</td>
            </tr>
            """

    html += """
        </table>
    </div>
    """
    return html

def build_shift_table(year, month, admin=False, manage=False, user=None):
    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"; end_date = f"{year}-{month:02d}-{last_day:02d}"
    conn = db_connect(); cur = conn.cursor()
    if manage:
        # 管理者の提出シフト管理では、確定済み・削り済みも含めて全員を表示する
        cur.execute("SELECT * FROM shifts WHERE date BETWEEN ? AND ? ORDER BY date ASC, start ASC", (start_date, end_date))
    elif admin:
        cur.execute("SELECT * FROM shifts WHERE date BETWEEN ? AND ? ORDER BY date ASC, start ASC", (start_date, end_date))
    else:
        cur.execute("SELECT * FROM shifts WHERE date BETWEEN ? AND ? AND confirmed = 1 AND cut = 0 ORDER BY date ASC, start ASC", (start_date, end_date))
    rows = cur.fetchall(); conn.close()
    help_grouped = get_help_applications(start_date, end_date)
    grouped = {}
    for r in rows:
        grouped.setdefault(r["date"], []).append(r)
    action = "/admin-shifts" if manage else "/admin-shift-table" if admin else "/shift-table"
    body = f"""
    <h2>{'提出シフト管理' if manage else 'シフト表一覧'}</h2>
    <form class="month-form" action="{action}" method="get">
        <div><label>年</label><input type="number" name="year" value="{year}"></div>
        <div><label>月</label><input type="number" name="month" value="{month}"></div>
        <button type="submit">表示</button>
    </form>
    {f'<a class="btn confirm" href="/admin-shift-new?year={year}&month={month}">＋ 新規シフト入力</a>' if manage else ''}
    """
    for d in range(1, last_day + 1):
        day = f"{year}-{month:02d}-{d:02d}"
        body += timeline_html(day, grouped.get(day, []), is_published(day), admin=admin, manage=manage, year=year, month=month, help_apps=help_grouped.get(day, []), viewer=user)
    if manage:
        body += build_pass_rate_table(start_date, end_date)
    body += '<a class="btn back" href="/admin" >戻る</a>' if manage or admin else '<a class="btn back" href="/portal">戻る</a>'
    return body


@app.get("/shift-table", response_class=HTMLResponse)
def shift_table(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user:
        return redirect("/")
    today = date.today(); year = year or today.year; month = month or today.month
    return layout("シフト表", build_shift_table(year, month, admin=False, user=user), user=user)


@app.get("/salary", response_class=HTMLResponse)
def salary(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user:
        return redirect("/")

    today = date.today()
    year = year or today.year
    month = month or today.month
    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT hourly_wage FROM users WHERE login_id = ?", (user["login_id"],))
    wage_row = cur.fetchone()
    hourly_wage = int(wage_row["hourly_wage"] or 0) if wage_row else 0

    # 削られたシフトは給与計算から除外。確定済みがあれば確定済みだけを優先、なければ提出済みを概算に使う。
    cur.execute("""
    SELECT * FROM shifts
    WHERE user_id = ? AND date BETWEEN ? AND ? AND cut = 0
    ORDER BY date ASC
    """, (user["login_id"], start_date, end_date))
    all_rows = cur.fetchall()
    confirmed_rows = [r for r in all_rows if int(r["confirmed"] or 0) == 1]
    rows_for_calc = confirmed_rows if confirmed_rows else all_rows
    conn.close()

    total_hours = sum(calc_hours(r["start"], r["end"]) for r in rows_for_calc)
    estimated_pay = int(total_hours * hourly_wage)
    calc_label = "確定シフトのみで計算" if confirmed_rows else "提出済みシフトで概算"

    details = ""
    if rows_for_calc:
        for r in rows_for_calc:
            h = calc_hours(r["start"], r["end"])
            details += f"""
            <tr>
                <td>{escape(r["date"])}</td>
                <td>{escape(r["start"] or "--:--")}-{escape(r["end"] or "--:--")}</td>
                <td>{h:.1f} h</td>
                <td>{int(h * hourly_wage):,} 円</td>
            </tr>
            """
    else:
        details = '<tr><td colspan="4">この月の計算対象シフトはありません。</td></tr>'

    body = f"""
    <h2>給与計算</h2>

    <div class="box">
        <form action="/salary-wage" method="post">
            <label>現在の時給</label>
            <input type="number" name="hourly_wage" value="{hourly_wage}" min="0" placeholder="例：1200">
            <button type="submit">時給を保存する</button>
        </form>
    </div>

    <form class="month-form" action="/salary" method="get">
        <div><label>年</label><input type="number" name="year" value="{year}"></div>
        <div><label>月</label><input type="number" name="month" value="{month}"></div>
        <button type="submit">表示</button>
    </form>

    <div class="summary">
        <div class="box"><div>計算対象時間</div><div class="summary-num">{total_hours:.1f} h</div></div>
        <div class="box"><div>概算給与</div><div class="summary-num">{estimated_pay:,} 円</div></div>
    </div>

    <div class="box">
        <b>{calc_label}</b><br>
        <span style="color:#b00020; font-weight:bold;">注意：</span>
        この給与計算は、登録された時給とシフト時間から計算した概算です。残業手当、深夜手当、交通費、控除、休憩時間、実際の勤怠打刻などは反映していないため、実際の給与とは異なる場合があります。
    </div>

    <div class="box">
        <h3>{year}年{month}月 明細</h3>
        <table>
            <tr><th>日付</th><th>時間</th><th>勤務時間</th><th>概算</th></tr>
            {details}
        </table>
    </div>

    <a class="btn back" href="/portal">戻る</a>
    """
    return layout("給与計算", body, user=user)


@app.post("/salary-wage")
def salary_wage(request: Request, hourly_wage: int = Form(...)):
    user = require_login(request)
    if not user:
        return redirect("/")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("UPDATE users SET hourly_wage = ? WHERE login_id = ?", (hourly_wage, user["login_id"]))
    conn.commit()
    conn.close()
    return redirect("/salary")


@app.get("/opinion", response_class=HTMLResponse)
def opinion(request: Request):
    user = require_login(request)
    if not user:
        return redirect("/")
    body = """
    <h2>意見箱</h2>
    <form action="/opinion-submit" method="post">
        <label>意見・要望</label><textarea name="message" rows="8" placeholder="改善してほしいことなど"></textarea>
        <button type="submit">送信する</button>
    </form>
    <a class="btn back" href="/portal">戻る</a>
    """
    return layout("意見箱", body, user=user)


@app.post("/opinion-submit")
def opinion_submit(request: Request, message: str = Form(...)):
    user = require_login(request)
    if not user:
        return redirect("/")
    conn = db_connect(); cur = conn.cursor()
    cur.execute("INSERT INTO opinions (user_id, name, message, created_at, status) VALUES (?, ?, ?, ?, 'pending')", (user["login_id"], user["name"], message, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit(); conn.close()
    return HTMLResponse("<script>alert('意見を送信しました！'); window.location='/portal';</script>")


def pending_opinion_count():
    try:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM opinions WHERE IFNULL(status,'pending')='pending'")
        n = cur.fetchone()[0]
        conn.close()
        return int(n or 0)
    except Exception:
        return 0

def pending_help_count():
    try:
        conn = db_connect()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM help_applications WHERE status = 'pending'")
        n = cur.fetchone()[0]
        conn.close()
        return int(n or 0)
    except Exception:
        return 0

@app.get("/admin", response_class=HTMLResponse)
def admin(request: Request):
    user = require_login(request)
    if not user:
        return redirect("/")
    if not is_admin_user(user):
        body = '<div class="box">管理者以外はこの画面を表示できません。</div><a class="btn back" href="/portal">戻る</a>'
        return layout("管理画面", body, user=user)

    pending_help = pending_help_count()
    pending_opinion = pending_opinion_count()
    help_badge = f"<span class='app-badge'>{pending_help}</span>" if pending_help > 0 else ""
    opinion_badge = f"<span class='app-badge'>{pending_opinion}</span>" if pending_opinion > 0 else ""
    notices = ""
    if pending_help > 0:
        notices += f"<div class='box notify-box'>ヘルプ応募が <b>{pending_help}</b> 件あります。</div>"
    if pending_opinion > 0:
        notices += f"<div class='box notify-box'>未確認の意見が <b>{pending_opinion}</b> 件あります。</div>"

    body = f"""
    <h2>管理画面</h2>
    {notices}
    <div class="grid">
        <div class="card" onclick="location.href='/admin-shifts'">提出シフト管理</div>
        <div class="card" onclick="location.href='/admin-submission-status'">シフト提出状況</div>
        <div class="card" onclick="location.href='/admin-shift-table'">シフト表確認</div>
        <div class="card" onclick="location.href='/admin-publish'">日付別公開設定</div>
        <div class="card" onclick="location.href='/admin-users'">従業員一覧</div>
        <div class="card badge-card" onclick="location.href='/admin-help'">ヘルプ応募確認 {help_badge}</div>
        <div class="card" onclick="location.href='/admin-labor'">人件費計算</div>
        <div class="card" onclick="location.href='/admin-export-excel'">Excel書き出し</div>
        <div class="card badge-card" onclick="location.href='/admin-opinions'">意見箱確認 {opinion_badge}</div>
        <div class="card" onclick="location.href='/admin-backups'">バックアップ</div>
        <div class="card" onclick="location.href='/portal'">戻る</div>
    </div>
    """
    return layout("管理画面", body, user=user)


@app.get("/admin-submission-status", response_class=HTMLResponse)
def admin_submission_status(request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    year, month, start_day, end_day, deadline, remaining_text = get_shift_period()
    start_date = f"{year}-{month:02d}-{start_day:02d}"
    end_date = f"{year}-{month:02d}-{end_day:02d}"

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT login_id, name, status FROM users ORDER BY CAST(login_id AS INTEGER) ASC, login_id ASC")
    employees = cur.fetchall()

    rows = []
    submitted_count = 0
    for emp in employees:
        cur.execute("""
        SELECT COUNT(*) AS c
        FROM shifts
        WHERE user_id = ? AND date BETWEEN ? AND ?
        """, (emp["login_id"], start_date, end_date))
        count = cur.fetchone()["c"]
        submitted = count > 0
        if submitted:
            submitted_count += 1
        rows.append((emp, count, submitted))
    conn.close()

    total = len(employees)
    percent = int(submitted_count / total * 100) if total else 0

    body = f"""
    <h2>シフト提出状況</h2>
    <div class="summary">
        <div class="box"><div>提出済み</div><div class="summary-num">{submitted_count}/{total}人</div></div>
        <div class="box"><div>提出率</div><div class="summary-num">{percent}%</div></div>
    </div>
    <div class="box">
        対象期間：<b>{year}年{month}月{start_day}日〜{end_day}日</b><br>
        提出締切：<b>{deadline.strftime('%Y/%m/%d')} 23:59</b><br>
        残り時間：<b>{remaining_text}</b>
    </div>
    <table>
        <tr><th>名前</th><th>ID</th><th>提出数</th><th>状態</th></tr>
    """

    for emp, count, submitted in rows:
        state = '<span class="status-ok">提出済み</span>' if submitted else '<span class="status-ng">未提出</span>'
        body += f"""
        <tr>
            <td>{escape(emp['name'])}</td>
            <td>{escape(emp['login_id'])}</td>
            <td>{count}</td>
            <td>{state}</td>
        </tr>
        """

    body += """
    </table>
    <a class="btn back" href="/admin">戻る</a>
    """
    return layout("シフト提出状況", body, user=user)


@app.get("/admin-shifts", response_class=HTMLResponse)
def admin_shifts(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    today = date.today(); year = year or today.year; month = month or today.month
    return layout("提出シフト管理", build_shift_table(year, month, manage=True, user=user), user=user, auto_scroll=False)



@app.get("/admin-shift-new", response_class=HTMLResponse)
def admin_shift_new(request: Request, year: int = None, month: int = None, day: str = ""):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    today = date.today()
    year = year or today.year
    month = month or today.month
    default_day = day or f"{year}-{month:02d}-{min(today.day, monthrange(year, month)[1]):02d}"

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT login_id, name, status FROM users ORDER BY CAST(login_id AS INTEGER) ASC, login_id ASC")
    users = cur.fetchall()
    conn.close()

    options = ""
    for u in users:
        label = f"{u['login_id']} {u['name']}"
        if "status" in u.keys() and u["status"]:
            label += f"（{u['status']}）"
        options += f'<option value="{escape(u["login_id"])}">{escape(label)}</option>'

    body = f"""
    <h2>新規シフト入力</h2>
    <div class="box">
        管理者が直接シフトを追加できます。追加したシフトは確定済みとして保存されます。
    </div>

    <form action="/admin-shift-new" method="post">
        <input type="hidden" name="return_year" value="{year}">
        <input type="hidden" name="return_month" value="{month}">

        <label>従業員</label>
        <select name="user_id" required>
            {options}
        </select>

        <label>日付</label>
        <input type="date" name="shift_date" value="{default_day}" required>

        <label>開始</label>
        <select name="start" required>{time_options()}</select>

        <label>終了</label>
        <select name="end" required>{time_options()}</select>

        <label>備考</label>
        <textarea name="memo" placeholder="例：急な追加、代理出勤、店長依頼など"></textarea>

        <button type="submit">シフトを追加する</button>
    </form>

    <a class="btn back" href="/admin-shifts?year={year}&month={month}#day-{default_day}">提出シフト管理へ戻る</a>
    <a class="btn back" href="/admin-shift-table?year={year}&month={month}#day-{default_day}">シフト表確認へ戻る</a>
    """
    return layout("新規シフト入力", body, user=user, auto_scroll=False)


@app.post("/admin-shift-new")
def admin_shift_new_submit(
    request: Request,
    user_id: str = Form(...),
    shift_date: str = Form(...),
    start: str = Form(...),
    end: str = Form(...),
    memo: str = Form(""),
    return_year: int = Form(0),
    return_month: int = Form(0)
):
    backup_before_change("before_manual_shift_add")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    start_h = parse_time_to_hour(start)
    end_h = parse_time_to_hour(end)
    if start_h is None or end_h is None or end_h <= start_h:
        return HTMLResponse("<script>alert('開始時間と終了時間を正しく入力してください。'); history.back();</script>")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE login_id = ?", (user_id,))
    target = cur.fetchone()
    if not target:
        conn.close()
        return HTMLResponse("<script>alert('従業員が見つかりません。'); history.back();</script>")

    cur.execute("""
    INSERT INTO shifts (user_id, name, date, start, end, limit_hour, memo, confirmed, cut, cut_memo)
    VALUES (?, ?, ?, ?, ?, ?, ?, 1, 0, '')
    """, (target["login_id"], target["name"], shift_date, start, end, "管理者追加", memo or "管理者追加"))

    conn.commit()
    conn.close()

    try:
        y = int(return_year) if return_year else int(shift_date.split("-")[0])
        m = int(return_month) if return_month else int(shift_date.split("-")[1])
    except Exception:
        y, m = date.today().year, date.today().month

    return redirect(f"/admin-shifts?year={y}&month={m}&no_auto_scroll=1#day-{shift_date}")


@app.get("/delete-shift-admin/{shift_id}")
def delete_shift_admin(shift_id: int, request: Request, year: int = None, month: int = None):
    backup_before_change("before_shift_delete")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT date FROM shifts WHERE id = ?", (shift_id,))
    row = cur.fetchone()
    target_date = row["date"] if row else ""
    cur.execute("DELETE FROM shifts WHERE id = ?", (shift_id,))
    conn.commit()
    conn.close()

    if target_date:
        try:
            y = year or int(target_date.split("-")[0])
            m = month or int(target_date.split("-")[1])
            return redirect(f"/admin-shifts?year={y}&month={m}&no_auto_scroll=1#day-{target_date}")
        except Exception:
            pass

    return redirect(f"/admin-shifts?year={year or date.today().year}&month={month or date.today().month}")


@app.get("/admin-shift-action/{shift_id}", response_class=HTMLResponse)
def admin_shift_action(shift_id: int, request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("SELECT * FROM shifts WHERE id = ?", (shift_id,)); r = cur.fetchone(); conn.close()
    if not r:
        return redirect("/admin-shifts")
    body = f"""
    <h2>シフト操作</h2>
    <div class="box">
        <b>{escape(r['name'])}</b><br>
        {escape(r['date'])}<br>
        {escape(r['start'] or '--:--')} - {escape(r['end'] or '--:--')}<br>
        希望：{escape(r['limit_hour'] or '')}<br>
        現在：{'削り済み' if int(r['cut'] or 0) == 1 else '未確定'}
    </div>
    <a class="btn confirm" href="/confirm-shift/{r['id']}?year={year or date.today().year}&month={month or date.today().month}">確定する</a>
    <a class="btn danger" href="/cut-shift/{r['id']}?year={year or date.today().year}&month={month or date.today().month}">削る</a>
    <a class="btn back" href="/admin-shifts?year={year or date.today().year}&month={month or date.today().month}">戻る</a>
    """
    return layout("シフト操作", body, user=user)


@app.get("/confirm-shift/{shift_id}")
def confirm_shift(shift_id: int, request: Request, year: int = None, month: int = None):
    backup_before_change("before_shift_confirm")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor()
    cur.execute("UPDATE shifts SET confirmed = 1, cut = 0 WHERE id = ?", (shift_id,))
    conn.commit(); conn.close()
    return redirect(f"/admin-shifts?year={year or date.today().year}&month={month or date.today().month}")


@app.get("/cut-shift/{shift_id}")
def cut_shift(shift_id: int, request: Request, year: int = None, month: int = None):
    backup_before_change("before_shift_cut")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor()
    cur.execute("UPDATE shifts SET cut = 1, confirmed = 0, cut_memo = ? WHERE id = ?", (datetime.now().strftime("%Y-%m-%d %H:%M"), shift_id))
    conn.commit(); conn.close()
    return redirect(f"/admin-shifts?year={year or date.today().year}&month={month or date.today().month}")


@app.get("/admin-shift-table", response_class=HTMLResponse)
def admin_shift_table(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    today = date.today(); year = year or today.year; month = month or today.month
    return layout("管理用シフト表", build_shift_table(year, month, admin=True, user=user), user=user)


@app.get("/admin-publish", response_class=HTMLResponse)
def admin_publish(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    today = date.today(); year = year or today.year; month = month or today.month
    last_day = monthrange(year, month)[1]
    body = f"""
    <h2>日付別公開設定</h2>
    <form class="month-form" action="/admin-publish" method="get">
        <div><label>年</label><input type="number" name="year" value="{year}"></div>
        <div><label>月</label><input type="number" name="month" value="{month}"></div>
        <button type="submit">表示</button>
    </form>
    <table class='publish-table'><tr><th>日付</th><th>状態</th><th>操作</th><th>メモ</th></tr>
    """
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    for d in range(1, last_day + 1):
        day = f"{year}-{month:02d}-{d:02d}"
        dt = date.fromisoformat(day)
        pub = is_published(day)
        state = "公開中" if pub else "非公開"
        action = "非公開にする" if pub else "公開する"
        val = 0 if pub else 1
        state_class = "publish-state state-on" if pub else "publish-state state-off"
        btn_class = "pub-off" if pub else "pub-on"
        memo_value = escape(get_day_memo(day))
        body += f"""<tr id="pub-row-{day}">
            <td>{month}/{d}({weekdays[dt.weekday()]})</td>
            <td id="pub-state-{day}" class="{state_class}">{state}</td>
            <td><a id="pub-btn-{day}" class="btn small-btn publish-btn {btn_class}" href="/set-publish/{day}/{val}" onclick="setPublishAjax('{day}', {val}); return false;">{action}</a></td>
            <td>
                <form action="/set-day-memo" method="post">
                    <input type="hidden" name="day" value="{day}">
                    <input type="hidden" name="year" value="{year}">
                    <input type="hidden" name="month" value="{month}">
                    <textarea name="memo" rows="2" placeholder="連絡事項">{memo_value}</textarea>
                    <button class="small-btn" type="submit">保存</button>
                </form>
            </td>
        </tr>"""
    body += "</table><a class='btn back' href='/admin'>戻る</a>"
    return layout("公開設定", body, user=user)


@app.get("/set-publish/{day}/{value}")
def set_publish(day: str, value: int, request: Request, ajax: int = 0):
    user = require_login(request)
    if not user or not is_admin_user(user):
        if ajax:
            return JSONResponse({"ok": False, "error": "not_admin"}, status_code=403)
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor()
    cur.execute("INSERT INTO published_days (date, published) VALUES (?, ?) ON CONFLICT(date) DO UPDATE SET published = excluded.published", (day, value))
    conn.commit(); conn.close()
    if ajax:
        return JSONResponse({"ok": True, "day": day, "published": value})
    y, m, _ = day.split("-")
    return redirect(f"/admin-publish?year={int(y)}&month={int(m)}#pub-row-{day}")



@app.post("/set-day-memo")
def set_day_memo(request: Request, day: str = Form(...), memo: str = Form(""), year: int = Form(None), month: int = Form(None)):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("INSERT INTO day_memos (date, memo) VALUES (?, ?) ON CONFLICT(date) DO UPDATE SET memo = excluded.memo", (day, memo.strip()))
    conn.commit()
    conn.close()
    y = year or int(day.split("-")[0])
    m = month or int(day.split("-")[1])
    return redirect(f"/admin-publish?year={y}&month={m}#pub-row-{day}")



@app.get("/help-apply", response_class=HTMLResponse)
def help_apply(request: Request, date: str, start: str, end: str, slot: int = 1, year: int = None, month: int = None):
    user = require_login(request)
    if not user:
        return redirect("/")
    try:
        parts = str(date).split("-")
        ret_y = int(year or parts[0])
        ret_m = int(month or parts[1])
    except Exception:
        today = date.today()
        ret_y = today.year
        ret_m = today.month

    body = f"""
    <h2>ヘルプ応募</h2>
    <div class="help-note">
        {escape(date)} のヘルプに応募します。勤務可能な時間を入力してください。
    </div>
    <form action="/help-apply-submit" method="post">
        <input type="hidden" name="return_year" value="{ret_y}">
        <input type="hidden" name="return_month" value="{ret_m}">
        <input type="hidden" name="return_date" value="{escape(date)}">
        <input type="hidden" name="date" value="{escape(date)}">
        <label>開始</label>
        <select name="start">{time_options(start)}</select>
        <label>終了</label>
        <select name="end">{time_options(end)}</select>
        <label>コメント（任意）</label>
        <textarea name="comment" placeholder="例：18時以降なら入れます"></textarea>
        <button type="submit">応募する</button>
    </form>
    <a class="btn back" href="/shift-table?year={ret_y}&month={ret_m}#day-{escape(date)}">戻る</a>
    """
    return layout("ヘルプ応募", body, user=user)




@app.post("/help-apply-submit")
def help_apply_submit(
    request: Request,
    date_value: str = Form(None),
    date: str = Form(None),
    start: str = Form(...),
    end: str = Form(...),
    comment: str = Form("")
):
    user = require_login(request)
    if not user:
        return redirect("/")

    target_date = date_value or date
    if not target_date:
        return HTMLResponse("<script>alert('日付が取得できませんでした。'); history.back();</script>")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
    INSERT INTO help_applications (user_id, name, date, start, end, comment, status, created_at)
    VALUES (?, ?, ?, ?, ?, ?, 'pending', ?)
    """, (user["login_id"], user["name"], target_date, start, end, comment, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

    try:
        y, m = target_date.split("-")[:2]
        return redirect(f"/shift-table?year={int(y)}&month={int(m)}&no_auto_scroll=1#day-{target_date}")
    except Exception:
        return redirect("/shift-table")



@app.get("/cancel-help/{app_id}")
def cancel_help(app_id: int, request: Request):
    user = require_login(request)
    if not user:
        return redirect("/")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT date FROM help_applications WHERE id = ? AND user_id = ?", (app_id, user["login_id"]))
    row = cur.fetchone()
    target_date = row["date"] if row else ""

    cur.execute("""
    UPDATE help_applications
    SET status = 'canceled'
    WHERE id = ? AND user_id = ? AND status = 'pending'
    """, (app_id, user["login_id"]))
    conn.commit()
    conn.close()

    if target_date:
        try:
            y, m = target_date.split("-")[:2]
            return redirect(f"/shift-table?year={int(y)}&month={int(m)}&no_auto_scroll=1#day-{target_date}")
        except Exception:
            pass

    return redirect("/shift-table")


@app.get("/admin-help", response_class=HTMLResponse)
def admin_help(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    today = date.today(); year = year or today.year; month = month or today.month
    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"; end_date = f"{year}-{month:02d}-{last_day:02d}"
    conn = db_connect(); cur = conn.cursor()
    cur.execute("SELECT * FROM help_applications WHERE date BETWEEN ? AND ? ORDER BY date ASC, start ASC", (start_date, end_date))
    rows = cur.fetchall(); conn.close()
    body = f"""
    <h2>ヘルプ応募確認</h2>
    <form class="month-form" action="/admin-help" method="get">
        <div><label>年</label><input type="number" name="year" value="{year}"></div>
        <div><label>月</label><input type="number" name="month" value="{month}"></div>
        <button type="submit">表示</button>
    </form>
    """
    if not rows:
        body += '<div class="box">ヘルプ応募はありません。</div>'
    for r in rows:
        status = "承認済み" if r["status"] == "approved" else "却下" if r["status"] == "rejected" else "未承認"
        actions = ""
        if r["status"] == "pending":
            actions = f"""
            <a class="btn confirm small-btn" href="/approve-help/{r['id']}?year={year}&month={month}">承認してシフト追加</a>
            <a class="btn danger small-btn" href="/reject-help/{r['id']}?year={year}&month={month}">却下</a>
            """
        body += f"""
        <div class="box">
            <b>{escape(r['name'])}</b><br>
            {escape(r['date'])}　{escape(r['start'])}-{escape(r['end'])}<br>
            状態：{status}<br>
            コメント：{escape(r['comment'] or '')}<br>
            {actions}
        </div>
        """
    body += '<a class="btn back" href="/admin">戻る</a>'
    return layout("ヘルプ応募確認", body, user=user, auto_scroll=False)


@app.get("/admin-help-action/{app_id}", response_class=HTMLResponse)
def admin_help_action(app_id: int, request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("SELECT * FROM help_applications WHERE id = ?", (app_id,)); r = cur.fetchone(); conn.close()
    if not r:
        return redirect("/admin-help")
    year = year or int(r["date"].split("-")[0]); month = month or int(r["date"].split("-")[1])
    body = f"""
    <h2>ヘルプ応募操作</h2>
    <div class="box">
        <b>{escape(r['name'])}</b><br>
        {escape(r['date'])}<br>
        {escape(r['start'])} - {escape(r['end'])}<br>
        コメント：{escape(r['comment'] or '')}<br>
    </div>
    <a class="btn confirm" href="/approve-help/{r['id']}?year={year}&month={month}">承認してシフト追加</a>
    <a class="btn danger" href="/reject-help/{r['id']}?year={year}&month={month}">却下</a>
    <a class="btn back" href="/admin-shifts?year={year}&month={month}">戻る</a>
    """
    return layout("ヘルプ応募操作", body, user=user, auto_scroll=False)


@app.get("/approve-help/{app_id}")
def approve_help(app_id: int, request: Request, year: int = None, month: int = None):
    backup_before_change("before_help_approve")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("SELECT * FROM help_applications WHERE id = ?", (app_id,)); r = cur.fetchone()
    if r:
        cur.execute("""
        INSERT INTO shifts (user_id, name, date, start, end, limit_hour, memo, confirmed, cut, cut_memo)
        VALUES (?, ?, ?, ?, ?, 'ヘルプ承認', ?, 1, 0, '')
        """, (r["user_id"], r["name"], r["date"], r["start"], r["end"], "ヘルプ応募から追加"))
        cur.execute("UPDATE help_applications SET status = 'approved' WHERE id = ?", (app_id,))
        conn.commit()
    conn.close()
    return redirect(f"/admin-help?year={year or date.today().year}&month={month or date.today().month}")


@app.get("/reject-help/{app_id}")
def reject_help(app_id: int, request: Request, year: int = None, month: int = None):
    backup_before_change("before_help_reject")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("UPDATE help_applications SET status = 'rejected' WHERE id = ?", (app_id,)); conn.commit(); conn.close()
    return redirect(f"/admin-help?year={year or date.today().year}&month={month or date.today().month}")


@app.get("/admin-users", response_class=HTMLResponse)
def admin_users(request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users ORDER BY id ASC")
    rows = cur.fetchall()
    conn.close()

    body = """
    <h2>従業員一覧</h2>
    <div class="box">
        従業員ごとの「編集する」ボタンから、ID・名前・パスワード・時給を変更できます。
    </div>
    """

    for r in rows:
        role = "管理者" if int(r["is_admin"] or 0) == 1 else "スタッフ"
        wage = int(r["hourly_wage"] or 0) if "hourly_wage" in r.keys() else 0
        emp_status = r["status"] if "status" in r.keys() and r["status"] else "アルバイト"

        admin_actions = ""
        if r["login_id"] != user["login_id"]:
            if int(r["is_admin"] or 0) == 1:
                admin_actions += f'<a class="btn danger small-btn" href="/revoke-admin/{r["id"]}" onclick="saveScrollNow()">権限剥奪</a>'
            else:
                admin_actions += f'<a class="btn small-btn" href="/grant-admin/{r["id"]}" onclick="saveScrollNow()">管理者にする</a>'
            admin_actions += f'<a class="btn danger small-btn" href="/delete-user-confirm/{r["id"]}" onclick="saveScrollNow()">従業員削除</a>'
        else:
            admin_actions += '<span style="font-weight:bold;color:#777;">自分自身は削除不可</span>'

        body += f"""
        <div class="box employee-row" id="user-{r['id']}">
            <div style="width:100%;">
                <div style="display:flex; justify-content:space-between; gap:12px; align-items:center; flex-wrap:wrap;">
                    <div>
                        <div style="font-size:20px;font-weight:900;">{escape(r['name'])}</div>
                        <div>ID：<b>{escape(r['login_id'])}</b></div>
                        <div>権限：<b>{role}</b></div>
                        <div>ステータス：<b>{escape(emp_status)}</b></div>
                        <div>時給：<b>{wage:,} 円</b></div>
                    </div>
                    <div style="display:flex; gap:8px; flex-wrap:wrap;">
                        <a class="btn small-btn" href="/admin-user-edit/{r['id']}" onclick="saveScrollNow()">編集する</a>
                        {admin_actions}
                    </div>
                </div>
            </div>
        </div>
        """

    body += """
    <script>
    function saveScrollNow(){
        sessionStorage.setItem('adminUsersScrollY', String(window.scrollY));
    }
    window.addEventListener('load', function(){
        const y = sessionStorage.getItem('adminUsersScrollY');
        if(y !== null){
            window.scrollTo(0, parseInt(y || '0', 10));
            sessionStorage.removeItem('adminUsersScrollY');
        }
    });
    </script>
    <a class="btn back" href="/admin">戻る</a>
    """
    return layout("従業員一覧", body, user=user)


@app.get("/admin-user-edit/{user_id}", response_class=HTMLResponse)
def admin_user_edit_page(user_id: int, request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    target = cur.fetchone()
    conn.close()

    if not target:
        return redirect("/admin-users")

    wage = int(target["hourly_wage"] or 0) if "hourly_wage" in target.keys() else 0
    emp_status = target["status"] if "status" in target.keys() and target["status"] else "アルバイト"
    role = "管理者" if int(target["is_admin"] or 0) == 1 else "スタッフ"

    body = f"""
    <h2>従業員情報編集</h2>

    <div class="box">
        <b>{escape(target['name'])}</b> さんの情報を編集します。<br>
        IDを変更すると、過去の提出シフト・ヘルプ応募・意見箱データも新しいIDへ移行されます。ステータスは「社員 / オフィシャルトレーナー / アルバイト」から選べます。
    </div>

    <form action="/admin-user-edit/{user_id}" method="post" onsubmit="return confirm('この従業員情報を保存しますか？\\nIDを変更した場合、過去データも新IDへ移行されます。');">
        <label>ログインID</label>
        <input name="login_id" value="{escape(target['login_id'])}" required>

        <label>名前</label>
        <input name="name" value="{escape(target['name'])}" required>

        <label>ステータス</label>
        <select name="status">
            <option value="社員" {"selected" if emp_status == "社員" else ""}>社員</option>
            <option value="オフィシャルトレーナー" {"selected" if emp_status == "オフィシャルトレーナー" else ""}>オフィシャルトレーナー</option>
            <option value="アルバイト" {"selected" if emp_status == "アルバイト" else ""}>アルバイト</option>
        </select>

        <label>新しいパスワード</label>
        <input type="text" name="password" value="" placeholder="変更しない場合は空欄">

        <label>時給</label>
        <input type="number" name="hourly_wage" value="{wage}" min="0" placeholder="例：1200">

        <div class="box">
            現在の権限：<b>{role}</b><br>
            権限変更は従業員一覧の「管理者にする / 権限剥奪」ボタンから行えます。
        </div>

        <button type="submit">保存する</button>
    </form>

    <a class="btn back" href="/admin-users">従業員一覧へ戻る</a>
    """
    return layout("従業員情報編集", body, user=user)


@app.post("/admin-user-edit/{user_id}")
def admin_user_edit_save(
    user_id: int,
    request: Request,
    login_id: str = Form(""),
    name: str = Form(""),
    password: str = Form(""),
    status: str = Form("アルバイト"),
    hourly_wage: int = Form(0)
):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    login_id = login_id.strip()
    name = name.strip()
    password = password.strip()
    status = status.strip()
    if status not in ["社員", "オフィシャルトレーナー", "アルバイト"]:
        status = "アルバイト"

    if not login_id or not name:
        return HTMLResponse("""
        <script>alert('IDと名前は必須です。'); history.back();</script>
        """)

    if hourly_wage < 0:
        hourly_wage = 0

    conn = db_connect()
    cur = conn.cursor()

    cur.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    target = cur.fetchone()
    if not target:
        conn.close()
        return redirect("/admin-users")

    old_login_id = target["login_id"]

    cur.execute("SELECT id FROM users WHERE login_id = ? AND id != ?", (login_id, user_id))
    duplicate = cur.fetchone()
    if duplicate:
        conn.close()
        return HTMLResponse("""
        <script>alert('そのIDはすでに使われています。'); history.back();</script>
        """)

    if password:
        cur.execute("""
        UPDATE users
        SET login_id = ?, name = ?, password = ?, status = ?, hourly_wage = ?
        WHERE id = ?
        """, (login_id, name, password, status, hourly_wage, user_id))
    else:
        cur.execute("""
        UPDATE users
        SET login_id = ?, name = ?, status = ?, hourly_wage = ?
        WHERE id = ?
        """, (login_id, name, status, hourly_wage, user_id))

    # 関連テーブルのIDと名前も更新
    related_tables = ["shifts", "help_applications", "opinions", "shift_history", "help_history"]
    for table in related_tables:
        try:
            if has_column(cur, table, "user_id"):
                cur.execute(f"UPDATE {table} SET user_id = ? WHERE user_id = ?", (login_id, old_login_id))
            if has_column(cur, table, "name"):
                cur.execute(f"UPDATE {table} SET name = ? WHERE user_id = ?", (name, login_id))
        except Exception:
            pass

    conn.commit()
    conn.close()

    res = redirect("/admin-users")
    if old_login_id == user["login_id"]:
        res.set_cookie("login_id", login_id, max_age=60 * 60 * 24 * 30)
    return res


@app.post("/update-user-info/{user_id}")
def update_user_info(
    user_id: int,
    request: Request,
    login_id: str = Form(""),
    name: str = Form(""),
    password: str = Form(""),
    hourly_wage: int = Form(0)
):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    login_id = login_id.strip()
    name = name.strip()
    password = password.strip()
    if hourly_wage < 0:
        hourly_wage = 0

    if not login_id or not name:
        return HTMLResponse("""
        <script>alert('IDと名前は必須です。'); history.back();</script>
        """)

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    target = cur.fetchone()
    if not target:
        conn.close()
        return redirect("/admin-users")

    old_login_id = target["login_id"]

    cur.execute("SELECT id FROM users WHERE login_id = ? AND id != ?", (login_id, user_id))
    duplicate = cur.fetchone()
    if duplicate:
        conn.close()
        return HTMLResponse("""
        <script>alert('そのIDはすでに使われています。'); history.back();</script>
        """)

    if password:
        cur.execute("UPDATE users SET login_id = ?, name = ?, password = ?, hourly_wage = ? WHERE id = ?",
                    (login_id, name, password, hourly_wage, user_id))
    else:
        cur.execute("UPDATE users SET login_id = ?, name = ?, hourly_wage = ? WHERE id = ?",
                    (login_id, name, hourly_wage, user_id))

    # IDや名前を変更した場合、関連テーブルもまとめて更新する
    for table in ["shifts", "help_applications", "opinions"]:
        try:
            if has_column(cur, table, "user_id"):
                cur.execute(f"UPDATE {table} SET user_id = ? WHERE user_id = ?", (login_id, old_login_id))
            if has_column(cur, table, "name"):
                cur.execute(f"UPDATE {table} SET name = ? WHERE user_id = ?", (name, login_id))
        except Exception:
            pass

    # もしシフト履歴系テーブルがある場合もできる範囲で更新
    for table in ["shift_history", "help_history"]:
        try:
            if has_column(cur, table, "user_id"):
                cur.execute(f"UPDATE {table} SET user_id = ? WHERE user_id = ?", (login_id, old_login_id))
            if has_column(cur, table, "name"):
                cur.execute(f"UPDATE {table} SET name = ? WHERE user_id = ?", (name, login_id))
        except Exception:
            pass

    conn.commit()
    conn.close()

    # 自分自身のIDを変更した場合はCookieも新IDに更新
    res = redirect("/admin-users")
    if old_login_id == user["login_id"]:
        res.set_cookie("login_id", login_id, max_age=60 * 60 * 24 * 30)
    return res


@app.post("/update-user-wage/{user_id}")
def update_user_wage(user_id: int, request: Request, hourly_wage: int = Form(...)):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    if hourly_wage < 0:
        hourly_wage = 0

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("UPDATE users SET hourly_wage = ? WHERE id = ?", (hourly_wage, user_id))
    conn.commit()
    conn.close()

    return redirect("/admin-users")


@app.get("/delete-user-confirm/{user_id}", response_class=HTMLResponse)
def delete_user_confirm(user_id: int, request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    target = cur.fetchone()
    conn.close()

    if not target:
        return redirect("/admin-users")

    if target["login_id"] == user["login_id"]:
        body = """
        <h2>削除できません</h2>
        <div class="box">自分自身のアカウントは削除できません。</div>
        <a class="btn back" href="/admin-users">戻る</a>
        """
        return layout("削除できません", body, user=user)

    body = f"""
    <h2>従業員削除の確認</h2>

    <div class="box" style="border:3px solid #e05252;background:#fff1f1;">
        <h3 style="color:#d40000;margin-top:0;">⚠ 警告 1</h3>
        <b>{escape(target['name'])}</b> さんを削除すると、この従業員はログインできなくなります。
    </div>

    <div class="box" style="border:3px solid #e05252;background:#fff1f1;">
        <h3 style="color:#d40000;margin-top:0;">⚠ 警告 2</h3>
        この従業員の提出シフト・確定シフト・ヘルプ応募は、すべてのシフト表から表示されなくなります。
        この操作は取り消せません。
    </div>

    <form action="/delete-user/{user_id}" method="post">
        <label>本当に削除する場合は、下に <b>削除</b> と入力してください</label>
        <input name="confirm_text" placeholder="削除">
        <button class="danger" type="submit">本当に削除する</button>
    </form>

    <a class="btn back" href="/admin-users">キャンセル</a>
    """
    return layout("従業員削除", body, user=user)


@app.post("/delete-user/{user_id}")
def delete_user(user_id: int, request: Request, confirm_text: str = Form("")):
    backup_before_change("before_user_delete")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    target = cur.fetchone()

    if not target:
        conn.close()
        return redirect("/admin-users")

    if target["login_id"] == user["login_id"]:
        conn.close()
        return redirect("/admin-users")

    if confirm_text.strip() != "削除":
        conn.close()
        return redirect(f"/delete-user-confirm/{user_id}")

    target_login_id = target["login_id"]

    # シフト表から完全に表示されないように、関連データを削除
    cur.execute("DELETE FROM shifts WHERE user_id = ?", (target_login_id,))
    cur.execute("DELETE FROM help_applications WHERE user_id = ?", (target_login_id,))
    cur.execute("DELETE FROM opinions WHERE user_id = ?", (target_login_id,))
    cur.execute("DELETE FROM users WHERE id = ?", (user_id,))

    conn.commit()
    conn.close()

    return redirect("/admin-users")


@app.get("/grant-admin/{user_id}")
def grant_admin(user_id: int, request: Request):
    backup_before_change("before_admin_change")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("UPDATE users SET is_admin = 1 WHERE id = ?", (user_id,)); conn.commit(); conn.close()
    return redirect("/admin-users")


@app.get("/revoke-admin/{user_id}")
def revoke_admin(user_id: int, request: Request):
    backup_before_change("before_admin_change")
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("UPDATE users SET is_admin = 0 WHERE id = ?", (user_id,)); conn.commit(); conn.close()
    return redirect("/admin-users")


@app.get("/admin-opinions", response_class=HTMLResponse)
def admin_opinions(request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect(); cur = conn.cursor(); cur.execute("SELECT * FROM opinions ORDER BY id DESC")
    rows = cur.fetchall(); conn.close()
    body = "<h2>意見箱確認</h2>"
    if not rows:
        body += '<div class="box">まだ意見はありません。</div>'
    for r in rows:
        status = r["status"] if "status" in r.keys() and r["status"] else "pending"
        if status == "done":
            status_html = '<span class="help-status-rejected">確認済み</span>'
            action = ""
        else:
            status_html = '<span class="help-status-pending">未確認</span>'
            action = f'<a class="btn confirm small-btn" href="/mark-opinion-read/{r["id"]}">確認済みにする</a>'
        body += f"""
        <div class='box'>
            <b>{escape(r['name'])}</b>　{status_html}<br>
            {escape(r['message'])}<br>
            <small>{escape(r['created_at'])}</small><br>
            {action}
        </div>
        """
    body += "<a class='btn back' href='/admin'>戻る</a>"
    return layout("意見箱確認", body, user=user)


@app.get("/mark-opinion-read/{opinion_id}")
def mark_opinion_read(opinion_id: int, request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("UPDATE opinions SET status = 'done' WHERE id = ?", (opinion_id,))
    conn.commit()
    conn.close()
    return redirect("/admin-opinions")


@app.get("/admin-labor", response_class=HTMLResponse)
def admin_labor(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    today = date.today()
    year = year or today.year
    month = month or today.month

    conn = db_connect()
    cur = conn.cursor()

    # その月の予算
    ym = f"{year}-{month:02d}"
    cur.execute("SELECT budget FROM labor_budgets WHERE ym = ?", (ym,))
    b = cur.fetchone()
    budget = int(b["budget"] or 0) if b else 0

    # 月別人件費 1〜12月
    monthly_rows = []
    for m in range(1, 13):
        last_day = monthrange(year, m)[1]
        start_date = f"{year}-{m:02d}-01"
        end_date = f"{year}-{m:02d}-{last_day:02d}"

        cur.execute("""
        SELECT s.*, u.hourly_wage
        FROM shifts s
        LEFT JOIN users u ON s.user_id = u.login_id
        WHERE s.date BETWEEN ? AND ?
          AND IFNULL(s.cut, 0) = 0
          AND IFNULL(s.confirmed, 0) = 1
        """, (start_date, end_date))
        shift_rows = cur.fetchall()

        total = 0
        for r in shift_rows:
            wage = int(r["hourly_wage"] or 0)
            total += int(calc_hours(r["start"], r["end"]) * wage)

        cur.execute("SELECT budget FROM labor_budgets WHERE ym = ?", (f"{year}-{m:02d}",))
        br = cur.fetchone()
        mbudget = int(br["budget"] or 0) if br else 0
        diff = mbudget - total
        monthly_rows.append({"month": m, "total": total, "budget": mbudget, "diff": diff})

    # 選択月の従業員別内訳
    last_day = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    cur.execute("""
    SELECT s.*, u.hourly_wage
    FROM shifts s
    LEFT JOIN users u ON s.user_id = u.login_id
    WHERE s.date BETWEEN ? AND ?
      AND IFNULL(s.cut, 0) = 0
      AND IFNULL(s.confirmed, 0) = 1
    ORDER BY s.name ASC, s.date ASC
    """, (start_date, end_date))
    selected_shifts = cur.fetchall()
    conn.close()

    employee = {}
    total_cost = 0
    total_hours = 0
    for r in selected_shifts:
        name = r["name"] or "未設定"
        wage = int(r["hourly_wage"] or 0)
        h = calc_hours(r["start"], r["end"])
        cost = int(h * wage)
        total_cost += cost
        total_hours += h
        if name not in employee:
            employee[name] = {"hours": 0, "cost": 0, "wage": wage}
        employee[name]["hours"] += h
        employee[name]["cost"] += cost
        employee[name]["wage"] = wage

    diff = budget - total_cost
    diff_text = f"+{diff:,} 円" if diff >= 0 else f"-{abs(diff):,} 円"
    diff_color = "#19a974" if diff >= 0 else "#e53935"

    max_value = max([x["total"] for x in monthly_rows] + [x["budget"] for x in monthly_rows] + [1])
    chart = ""
    for x in monthly_rows:
        total_w = int((x["total"] / max_value) * 100)
        budget_w = int((x["budget"] / max_value) * 100) if x["budget"] else 0
        d = x["diff"]
        d_text = f"+{d:,}" if d >= 0 else f"-{abs(d):,}"
        d_color = "#19a974" if d >= 0 else "#e53935"
        chart += f"""
        <div class="labor-chart-row">
            <div class="labor-month">{x['month']}月</div>
            <div class="labor-bars">
                <div class="labor-bar total" style="width:{total_w}%">実績 {x['total']:,}</div>
                <div class="labor-bar budget" style="width:{budget_w}%">予算 {x['budget']:,}</div>
            </div>
            <div class="labor-diff" style="color:{d_color};">{d_text}</div>
        </div>
        """

    emp_rows = ""
    if employee:
        for name, v in employee.items():
            emp_rows += f"""
            <tr>
                <td>{escape(name)}</td>
                <td>{v['hours']:.1f} h</td>
                <td>{int(v['wage']):,} 円</td>
                <td>{int(v['cost']):,} 円</td>
            </tr>
            """
    else:
        emp_rows = '<tr><td colspan="4">この月の確定シフトはありません。</td></tr>'

    extra_style = """
    <style>
        .labor-summary { display:grid; grid-template-columns:1fr 1fr; gap:12px; }
        .labor-card { background:white; border-radius:18px; padding:18px; box-shadow:0 6px 18px rgba(0,0,0,.08); }
        .labor-number { font-size:28px; font-weight:900; margin-top:8px; }
        .labor-chart-row { display:grid; grid-template-columns:48px 1fr 82px; gap:10px; align-items:center; margin:12px 0; }
        .labor-month { font-weight:800; }
        .labor-bars { background:#f2f5f0; border-radius:12px; padding:6px; }
        .labor-bar { height:24px; border-radius:999px; margin:4px 0; color:white; font-size:12px; font-weight:800; line-height:24px; padding-left:8px; box-sizing:border-box; white-space:nowrap; min-width:48px; }
        .labor-bar.total { background:#8CC63F; }
        .labor-bar.budget { background:#333; }
        .labor-diff { font-weight:900; text-align:right; }
        @media(max-width:600px){ .labor-summary{grid-template-columns:1fr;} .labor-chart-row{grid-template-columns:40px 1fr 70px;} }
    </style>
    """

    body = f"""
    {extra_style}
    <h2>人件費計算</h2>

    <form class="month-form" action="/admin-labor" method="get">
        <div><label>年</label><input type="number" name="year" value="{year}"></div>
        <div><label>月</label><input type="number" name="month" value="{month}"></div>
        <button type="submit">表示</button>
    </form>

    <div class="box">
        <form action="/admin-labor-budget" method="post">
            <input type="hidden" name="year" value="{year}">
            <input type="hidden" name="month" value="{month}">
            <label>{year}年{month}月の予算</label>
            <input type="number" name="budget" value="{budget}" min="0" placeholder="例：500000">
            <button type="submit">予算を保存</button>
        </form>
    </div>

    <div class="labor-summary">
        <div class="labor-card"><div>総人件費</div><div class="labor-number">{total_cost:,} 円</div></div>
        <div class="labor-card"><div>総勤務時間</div><div class="labor-number">{total_hours:.1f} h</div></div>
        <div class="labor-card"><div>予算</div><div class="labor-number">{budget:,} 円</div></div>
        <div class="labor-card"><div>予算差額</div><div class="labor-number" style="color:{diff_color};">{diff_text}</div></div>
    </div>

    <div class="box" style="margin-top:16px;">
        <b>注意：</b>この人件費は、従業員の登録時給と確定シフト時間から計算した概算です。深夜手当、残業、休憩、交通費、控除、実際の勤怠打刻は反映していません。
    </div>

    <div class="box">
        <h3>{year}年 月別グラフ</h3>
        {chart}
    </div>

    <div class="box">
        <h3>{year}年{month}月 従業員別内訳</h3>
        <table>
            <tr><th>名前</th><th>時間</th><th>時給</th><th>人件費</th></tr>
            {emp_rows}
        </table>
    </div>

    <a class="btn back" href="/admin">戻る</a>
    """
    return layout("人件費計算", body, user=user)


@app.post("/admin-labor-budget")
def admin_labor_budget(request: Request, year: int = Form(...), month: int = Form(...), budget: int = Form(...)):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")
    ym = f"{year}-{month:02d}"
    conn = db_connect()
    cur = conn.cursor()
    cur.execute("""
    INSERT INTO labor_budgets (ym, budget)
    VALUES (?, ?)
    ON CONFLICT(ym) DO UPDATE SET budget = excluded.budget
    """, (ym, budget))
    conn.commit()
    conn.close()
    return redirect(f"/admin-labor?year={year}&month={month}")

@app.get("/admin-export-excel", response_class=HTMLResponse)
def admin_export_excel(request: Request, year: int = None, month: int = None):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    today = date.today()
    year = year or today.year
    month = month or today.month
    last_day = monthrange(year, month)[1]

    body = f"""
    <h2>Excel書き出し</h2>
    <div class="box">
        <p>印刷しやすいように、1日〜15日・16日〜月末・月全体を分けてExcel出力できます。</p>
        <form class="month-form" action="/admin-export-excel" method="get">
            <div><label>年</label><input type="number" name="year" value="{year}"></div>
            <div><label>月</label><input type="number" name="month" value="{month}"></div>
            <button type="submit">表示月を変更</button>
        </form>
    </div>

    <div class="box">
        <h3>{year}年{month}月 Excel出力</h3>
        <a class="btn confirm" href="/admin-export-excel-download?year={year}&month={month}&period=first">前半（1日〜15日）をダウンロード</a>
        <a class="btn confirm" href="/admin-export-excel-download?year={year}&month={month}&period=second">後半（16日〜{last_day}日）をダウンロード</a>
        <a class="btn" href="/admin-export-excel-download?year={year}&month={month}&period=all">月全体をダウンロード</a>
    </div>

    <div class="box">
        <b>出力内容</b><br>
        ・従業員一覧（ID順）<br>
        ・期間内の確定シフト<br>
        ・朝/夜などの人員不足メモ<br>
        ・日付メモ<br>
        ・集計シート（従業員別の合計時間）<br>
        ・セルサイズはそのまま、文字は見やすく大きめ
    </div>

    <a class="btn back" href="/admin">戻る</a>
    """
    return layout("Excel書き出し", body, user=user)


def excel_time_text(start, end):
    if not start or not end:
        return ""
    return f"{start.replace(':00', '')}-{end.replace(':00', '')}"


def shortage_text_for_excel(shifts):
    segments = help_segments(shifts)
    if not segments:
        return ""

    middle_parts = []
    morning_deficit = 0
    night_deficit = 0

    for hs, he, deficit in segments:
        if deficit <= 0:
            continue

        # 9:00-10:00 にかかる不足は「朝○人」にまとめる
        if hs < START_HOUR + 1 and he > START_HOUR:
            morning_deficit = max(morning_deficit, deficit)
            # 朝枠を除いた後ろ側があるなら中間不足として残す
            if he > START_HOUR + 1:
                s = hour_to_str(START_HOUR + 1).replace(":00", "")
                e = hour_to_str(he).replace(":00", "")
                middle_parts.append(f"{s}-{e} {deficit}人")
            continue

        # 21:00-22:00 にかかる不足は「夜○人」に1つだけまとめる
        if hs < END_HOUR and he > END_HOUR - 1:
            night_deficit = max(night_deficit, deficit)
            # 夜枠より前に不足が伸びていたら中間不足として残す
            if hs < END_HOUR - 1:
                s = hour_to_str(hs).replace(":00", "")
                e = hour_to_str(END_HOUR - 1).replace(":00", "")
                middle_parts.append(f"{s}-{e} {deficit}人")
            continue

        s = hour_to_str(hs).replace(":00", "")
        e = hour_to_str(he).replace(":00", "")
        middle_parts.append(f"{s}-{e} {deficit}人")

    parts = []
    if morning_deficit > 0:
        parts.append(f"朝{morning_deficit}人")
    parts.extend(middle_parts)
    if night_deficit > 0:
        parts.append(f"夜{night_deficit}人")

    return " / ".join(parts)


@app.get("/admin-export-excel-download")
def admin_export_excel_download(request: Request, year: int = None, month: int = None, period: str = "all"):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return HTMLResponse("""
        <h2>openpyxl が必要です</h2>
        <p>ターミナルで以下を実行してから、もう一度ダウンロードしてください。</p>
        <pre>python3 -m pip install openpyxl</pre>
        <a href="/admin-export-excel">戻る</a>
        """)

    today = date.today()
    year = year or today.year
    month = month or today.month
    last_day = monthrange(year, month)[1]

    if period == "first":
        start_day = 1
        end_day = min(15, last_day)
        period_label = "前半"
    elif period == "second":
        start_day = 16
        end_day = last_day
        period_label = "後半"
    else:
        start_day = 1
        end_day = last_day
        period_label = "月全体"

    start_date = f"{year}-{month:02d}-{start_day:02d}"
    end_date = f"{year}-{month:02d}-{end_day:02d}"
    day_numbers = list(range(start_day, end_day + 1))

    conn = db_connect()
    cur = conn.cursor()
    cur.execute("SELECT login_id, name, status FROM users ORDER BY CAST(login_id AS INTEGER) ASC, login_id ASC")
    employees = cur.fetchall()

    cur.execute("""
        SELECT * FROM shifts
        WHERE date BETWEEN ? AND ? AND confirmed = 1 AND cut = 0
        ORDER BY date ASC, start ASC
    """, (start_date, end_date))
    shifts = cur.fetchall()

    cur.execute("SELECT date, memo FROM day_memos WHERE date BETWEEN ? AND ?", (start_date, end_date))
    memo_rows = cur.fetchall()
    conn.close()

    memo_map = {r["date"]: r["memo"] for r in memo_rows}
    shifts_by_user_day = {}
    shifts_by_day = {}
    for r in shifts:
        shifts_by_user_day.setdefault((r["user_id"], r["date"]), []).append(r)
        shifts_by_day.setdefault(r["date"], []).append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月{period_label}"

    # colors
    green = "8CC63F"
    light_green = "D9EAD3"
    cream = "FFF2CC"
    pink = "F4CCCC"
    blue = "C9DAF8"
    purple = "D9D2E9"
    yellow = "FFD966"
    white = "FFFFFF"

    thin = Side(style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # title row / shortage row
    ws.cell(1, 1, "人数不足")
    ws.cell(1, 2, "")
    ws.cell(2, 1, "No")
    ws.cell(2, 2, "従業員")

    for idx, d in enumerate(day_numbers, start=1):
        day = f"{year}-{month:02d}-{d:02d}"
        dt = date(year, month, d)
        youbi = ["月", "火", "水", "木", "金", "土", "日"][dt.weekday()]
        col = idx + 2
        ws.cell(1, col, shortage_text_for_excel(shifts_by_day.get(day, [])))
        ws.cell(2, col, f"{d}日（{youbi}）")

    max_col = len(day_numbers) + 2

    # headers
    for row in [1, 2]:
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=13)
            if row == 1:
                cell.fill = PatternFill("solid", fgColor=purple if col <= 2 else yellow)
            else:
                cell.fill = PatternFill("solid", fgColor=light_green)

    # employee rows
    first_emp_row = 3
    summary = {}
    for i, emp in enumerate(employees, start=1):
        row = first_emp_row + i - 1
        emp_name = emp["name"] or emp["login_id"]
        ws.cell(row, 1, i)
        ws.cell(row, 2, f"{emp['login_id']} {emp_name}")
        summary[emp["login_id"]] = {
            "name": emp_name,
            "status": emp["status"] if "status" in emp.keys() and emp["status"] else "",
            "hours": 0.0,
            "days": 0
        }

        for base_col in [1, 2]:
            ws.cell(row, base_col).fill = PatternFill("solid", fgColor=cream)

        for idx, d in enumerate(day_numbers, start=1):
            col = idx + 2
            day = f"{year}-{month:02d}-{d:02d}"
            day_shifts = shifts_by_user_day.get((emp["login_id"], day), [])
            value = "\n".join(excel_time_text(r["start"], r["end"]) for r in day_shifts if r["start"] and r["end"])
            cell = ws.cell(row, col, value)

            if day_shifts:
                summary[emp["login_id"]]["days"] += 1
                summary[emp["login_id"]]["hours"] += sum(calc_hours(r["start"], r["end"]) for r in day_shifts)

            if value:
                # 朝昼はピンク、夕方以降は青
                first_start = parse_time_to_hour(day_shifts[0]["start"])
                fill = pink if first_start is not None and first_start < 15 else blue
                cell.fill = PatternFill("solid", fgColor=fill)
            else:
                cell.fill = PatternFill("solid", fgColor=white)

            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, max_col + 1):
            ws.cell(row, col).border = border
            ws.cell(row, col).font = Font(size=12)

    # memo row
    memo_row = first_emp_row + len(employees)
    ws.cell(memo_row, 1, "")
    ws.cell(memo_row, 2, "メモ")
    for col in range(1, max_col + 1):
        ws.cell(memo_row, col).fill = PatternFill("solid", fgColor=yellow)
        ws.cell(memo_row, col).border = border
        ws.cell(memo_row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(memo_row, col).font = Font(bold=True if col <= 2 else False, size=12)

    for idx, d in enumerate(day_numbers, start=1):
        day = f"{year}-{month:02d}-{d:02d}"
        ws.cell(memo_row, idx + 2, memo_map.get(day, ""))

    # formatting - セルサイズは従来と同じ
    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    for col in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    for row in range(3, memo_row + 1):
        ws.row_dimensions[row].height = 24

    ws.sheet_view.showGridLines = False

    # 集計シート
    summary_ws = wb.create_sheet("集計")
    summary_ws.cell(1, 1, f"{year}年{month}月 {period_label} 集計")
    summary_ws.cell(2, 1, "ID")
    summary_ws.cell(2, 2, "名前")
    summary_ws.cell(2, 3, "ステータス")
    summary_ws.cell(2, 4, "出勤日数")
    summary_ws.cell(2, 5, "合計時間")

    for col in range(1, 6):
        c = summary_ws.cell(2, col)
        c.font = Font(bold=True, size=13)
        c.fill = PatternFill("solid", fgColor=light_green)
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    for emp in employees:
        s = summary.get(emp["login_id"], {})
        summary_ws.cell(row, 1, emp["login_id"])
        summary_ws.cell(row, 2, s.get("name", ""))
        summary_ws.cell(row, 3, s.get("status", ""))
        summary_ws.cell(row, 4, s.get("days", 0))
        summary_ws.cell(row, 5, round(s.get("hours", 0.0), 2))
        for col in range(1, 6):
            c = summary_ws.cell(row, col)
            c.font = Font(size=12)
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    summary_ws.column_dimensions["A"].width = 12
    summary_ws.column_dimensions["B"].width = 20
    summary_ws.column_dimensions["C"].width = 20
    summary_ws.column_dimensions["D"].width = 12
    summary_ws.column_dimensions["E"].width = 12
    summary_ws.sheet_view.showGridLines = False

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    period_name = "first" if period == "first" else "second" if period == "second" else "all"
    filename = f"FE_shift_{year}_{month:02d}_{period_name}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )




@app.get("/admin-backups", response_class=HTMLResponse)
def admin_backups(request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    BACKUP_DIR.mkdir(exist_ok=True)
    backups = sorted(BACKUP_DIR.glob("fe_portal_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)

    rows = ""
    if not backups:
        rows = "<div class='box'>バックアップはまだありません。</div>"
    else:
        for b in backups:
            t = datetime.fromtimestamp(b.stat().st_mtime).strftime("%Y/%m/%d %H:%M:%S")
            size = round(b.stat().st_size / 1024, 1)
            rows += f"""
            <div class="box">
                <b>{escape(b.name)}</b><br>
                作成日時：{t}<br>
                サイズ：{size} KB
            </div>
            """

    json_info = "なし"
    if USER_JSON_BACKUP.exists():
        json_info = datetime.fromtimestamp(USER_JSON_BACKUP.stat().st_mtime).strftime("%Y/%m/%d %H:%M:%S")

    body = f"""
    <h2>バックアップ</h2>

    <div class="box">
        <b>最新DBバックアップ</b><br>
        {escape(latest_backup_info())}<br><br>
        <b>従業員JSONバックアップ</b><br>
        {escape(json_info)}
    </div>

    <form action="/admin-backup-create" method="post">
        <button type="submit">今すぐバックアップ作成</button>
    </form>

    <div class="box">
        自動バックアップは最新{BACKUP_KEEP}個まで保存されます。<br>
        従業員情報は users_backup.json にも保存されます。
    </div>

    {rows}

    <a class="btn back" href="/admin">管理画面へ戻る</a>
    """
    return layout("バックアップ", body, user=user)


@app.post("/admin-backup-create")
def admin_backup_create(request: Request):
    user = require_login(request)
    if not user or not is_admin_user(user):
        return redirect("/portal")

    create_backup("manual")
    save_users_json()
    return redirect("/admin-backups")
